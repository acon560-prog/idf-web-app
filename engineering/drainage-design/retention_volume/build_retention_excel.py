#!/usr/bin/env python3
"""
Build Excel: required retention volume upstream of controlling Ø900 culvert.

System (series):
  Ø1500 (in) → storage zone → Ø900 (control) → Ø1200 (downstream) → outlet

Method A: constant outlet capacity Qout = Qcap_900 (default 1.77 m³/s)
  V(t) = integral max(Qin - Qout, 0) dt   [trapezoidal]
  Vmax = max cumulative storage (storage cannot go negative)

Method B (optional): Qout = f(H) rating for Ø900, with simple level-pool
  using an assumed constant pond area (editable).
"""

from __future__ import annotations

import math
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HERE = Path(__file__).resolve().parent
OUT = HERE / "Volume_Retention_Ponceau_900.xlsx"

# Measured hydrograph (min, m³/s) — crossings Qin=Qcap inserted at build time
HYDRO_RAW = [
    (0.0, 0.778),
    (5.0, 1.170),
    (10.0, 1.857),
    (15.0, 3.202),
    (20.0, 6.309),
    (22.5, 9.600),
    (27.5, 5.846),
    (32.5, 3.857),
    (37.5, 2.689),
    (42.5, 1.953),
    (47.5, 1.462),
    (52.5, 1.121),
    (57.5, 0.875),
]

# Ø900 geometry + Manning (full pipe) — user: n=0.013, Q = capacité pleine section
N_MANNING = 0.013
D_900_M = 0.9
L_900 = 50.65
Z_US_900 = 34.88
Z_DS_900 = 34.45
S0_900 = (Z_US_900 - Z_DS_900) / L_900


def q_full_manning(D: float, S0: float, n: float) -> float:
    A = math.pi * D**2 / 4.0
    R = D / 4.0
    return (1.0 / n) * A * (R ** (2.0 / 3.0)) * (S0**0.5)


Q_FULL_900 = q_full_manning(D_900_M, S0_900, N_MANNING)  # ≈ 1.668 m³/s
# Prior office value (~1.77) was same method, slight rounding / S0 difference
QCAP_900_PRIOR = 1.77
QCAP_900 = round(Q_FULL_900, 3)  # default used in routing
QCAP_1200 = 4.93
QPEAK_IN = 9.60


def insert_qcap_crossings(hydro: list[tuple[float, float]], qcap: float) -> list[tuple[float, float]]:
    """Insert linearly interpolated times where Qin crosses qcap (up and down)."""
    out: list[tuple[float, float]] = [hydro[0]]
    for (t0, q0), (t1, q1) in zip(hydro, hydro[1:]):
        crossed_up = q0 < qcap <= q1
        crossed_dn = q0 > qcap >= q1
        if crossed_up or crossed_dn:
            if abs(q1 - q0) > 1e-12:
                t_x = t0 + (qcap - q0) / (q1 - q0) * (t1 - t0)
                out.append((round(t_x, 2), qcap))
        out.append((t1, q1))
    return out


HYDRO = insert_qcap_crossings(HYDRO_RAW, QCAP_900)

HEADER = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="0F5C5C")
YELLOW = PatternFill("solid", fgColor="FFF59D")
GREEN = PatternFill("solid", fgColor="D1FAE5")
ORANGE = PatternFill("solid", fgColor="FCE4D6")
BLUE = PatternFill("solid", fgColor="DBEAFE")
THIN = Border(
    left=Side(style="thin", color="CBD5E1"),
    right=Side(style="thin", color="CBD5E1"),
    top=Side(style="thin", color="CBD5E1"),
    bottom=Side(style="thin", color="CBD5E1"),
)


def style_header(ws, row: int, ncols: int) -> None:
    for c in range(1, ncols + 1):
        cell = ws.cell(row, c)
        cell.font = HEADER
        cell.fill = HEADER_FILL
        cell.border = THIN


def build() -> Path:
    wb = Workbook()

    # ---------- Parametres ----------
    ws = wb.active
    ws.title = "Parametres"
    ws["A1"] = "Volume de rétention — ponceau Ø900 contrôlant"
    ws["A1"].font = Font(bold=True, size=14, color="0F5C5C")
    ws.merge_cells("A1:F1")

    ws["A3"] = "Configuration (série)"
    ws["A3"].font = Font(bold=True)
    layout = [
        ("Ouvrage", "Rôle", "D (mm)", "Radier amont (m)", "Radier aval (m)", "Longueur (m)", "S0 (-)", "Q capacité donnée (m³/s)"),
        ("Ø1500", "Entrée / amont", 1500, 36.10, 35.05, 57.00, (36.10 - 35.05) / 57.00, QPEAK_IN),
        ("Ø900", "Contrôle sortie rétention", 900, 34.88, 34.45, 50.65, (34.88 - 34.45) / 50.65, QCAP_900),
        ("Ø1200", "Aval du Ø900", 1200, 33.60, 33.10, 31.30, (33.60 - 33.10) / 31.30, QCAP_1200),
    ]
    for j, h in enumerate(layout[0], start=1):
        ws.cell(4, j, h)
    style_header(ws, 4, 8)
    for i, row in enumerate(layout[1:], start=5):
        for j, v in enumerate(row, start=1):
            cell = ws.cell(i, j, v)
            cell.border = THIN
            if j in (4, 5, 6, 7, 8) and isinstance(v, float):
                cell.number_format = "0.000" if j != 7 else "0.00000"
        if row[0] == "Ø900":
            for j in range(1, 9):
                ws.cell(i, j).fill = ORANGE

    ws["A9"] = (
        "Chaîne: Ø1500 → zone de rétention → Ø900 (contrôle) → Ø1200 → exutoire. "
        "Le Ø1200 ne peut pas évacuer plus que ce que le Ø900 lui livre."
    )
    ws.merge_cells("A9:H10")
    ws["A9"].alignment = Alignment(wrap_text=True, vertical="top")

    # --- Manning full-pipe capacity check ---
    ws["A12"] = "Capacité Ø900 — Manning pleine section (n = 0,013)"
    ws["A12"].font = Font(bold=True)
    ws["A13"] = "n Manning (-)"
    ws["B13"] = N_MANNING
    ws["B13"].fill = YELLOW
    ws["B13"].border = THIN
    ws["B13"].number_format = "0.000"
    ws["C13"] = "Fourni (béton typique ~0,012–0,015)"
    ws["A14"] = "D (m)"
    ws["B14"] = D_900_M
    ws["B14"].number_format = "0.00"
    ws["A15"] = "S0 = (Zamont−Zaval)/L"
    ws["B15"] = S0_900
    ws["B15"].number_format = "0.00000"
    ws["C15"] = f"= ({Z_US_900} − {Z_DS_900}) / {L_900}"
    ws["A16"] = "A = πD²/4 (m²)"
    ws["B16"] = math.pi * D_900_M**2 / 4
    ws["B16"].number_format = "0.000"
    ws["A17"] = "R = D/4 (m)"
    ws["B17"] = D_900_M / 4
    ws["B17"].number_format = "0.000"
    ws["A18"] = "Q_plein = (1/n)·A·R^(2/3)·√S0"
    ws["B18"] = Q_FULL_900
    ws["B18"].fill = GREEN
    ws["B18"].border = THIN
    ws["B18"].number_format = "0.000"
    ws["C18"] = "m³/s — capacité pleine section (hypothèse actuelle)"
    ws["A19"] = "Q utilisé avant (m³/s)"
    ws["B19"] = QCAP_900_PRIOR
    ws["B19"].number_format = "0.00"
    ws["C19"] = "Ancienne valeur bureau (~6% plus haute; même méthode)"

    ws["A21"] = "Entrées de calcul rétention (jaune)"
    ws["A21"].font = Font(bold=True)
    ws["A22"] = "Qout_cap_900 (m³/s)"
    ws["B22"] = QCAP_900
    ws["B22"].fill = YELLOW
    ws["B22"].border = THIN
    ws["B22"].number_format = "0.000"
    ws["C22"] = "Par défaut = Q_plein Manning. Mettre 1.77 pour retrouver l'ancien calcul."
    ws["A23"] = "Facteur_securite (-)"
    ws["B23"] = 1.20
    ws["B23"].fill = YELLOW
    ws["B23"].border = THIN
    ws["B23"].number_format = "0.00"
    ws["C23"] = "Marge sur Vmax (ex. 1.20 = +20%)"
    ws["A24"] = "Aire_bassin_m2"
    ws["B24"] = 2500
    ws["B24"].fill = YELLOW
    ws["B24"].border = THIN
    ws["B24"].number_format = "0"
    ws["C24"] = "Aire plan d'eau approx. pour H = V/A (Méthode B)"

    ws["A26"] = "Résultats (formules — feuille Calcul_A)"
    ws["A26"].font = Font(bold=True)
    ws["A27"] = "Vmax (m³)"
    ws["B27"] = "=Calcul_A!B3"
    ws["B27"].fill = GREEN
    ws["B27"].number_format = "0"
    ws["A28"] = "V_dimensionnement = Vmax × facteur (m³)"
    ws["B28"] = "=B27*B23"
    ws["B28"].fill = GREEN
    ws["B28"].number_format = "0"
    ws["A29"] = "Hmax approx (m) = Vmax / Aire"
    ws["B29"] = "=IF(B24>0,B27/B24,0)"
    ws["B29"].fill = BLUE
    ws["B29"].number_format = "0.00"
    ws["C29"] = "Estimation grossière si l'aire de bassin est constante"

    ws["A31"] = "Inlet control vs outlet control — qu'est-ce que ça veut dire?"
    ws["A31"].font = Font(bold=True)
    ic_oc = [
        "• Inlet control (contrôle à l'entrée): le débit est limité par l'entrée du ponceau (forme d'entrée, HW/D).",
        "  Le tuyau aval « n'aspire » pas assez pour influencer — Q dépend surtout de la charge amont H, pas de L ni de n.",
        "• Outlet control (contrôle à la sortie / frottement): le débit dépend de la longueur, de n, de la pente et de la charge",
        "  (énergie amont → aval). Manning pleine section est une hypothèse de type outlet/frottement à section pleine.",
        f"• Ici L/D = {L_900/D_900_M:.0f} (tuyau assez long) → le contrôle par frottement (outlet) est souvent plausible,",
        "  mais sans analyse FHWA/HY-8 on ne peut pas l'affirmer. Pour dimensionner la rétention, Q_plein constant est",
        "  en général prudent (si la charge monte, le débit réel peut être un peu plus élevé → Vmax un peu plus bas).",
        "• On n'a pas besoin de trancher IC/OC pour utiliser ce classeur: garder Qout = Q_plein (ou 1.77) suffit pour une 1re estimation.",
    ]
    for i, line in enumerate(ic_oc, start=32):
        ws.cell(i, 1, line)
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=8)

    ws["A41"] = "Notes"
    ws["A41"].font = Font(bold=True)
    notes = [
        "• Méthode A: Qout fixe = Q_plein Manning Ø900 (n=0,013). Pas un jugement IC/OC complet.",
        "• Le stockage commence quand Qin dépasse Qout; Vmax = max du volume stocké (pas (Qp−Qout)×durée).",
        "• Méthode B: Qout = f(H) depuis feuille Courbe_QH_900 (FHWA inlet+outlet).",
        "• Voir feuille Fichiers_lies pour les anciens Excel HY-8 / ponceau+fossé.",
        "• Vérifier que le radier amont Ø900 (34.88) est bien le fond de la zone de rétention étudiée.",
    ]
    for i, line in enumerate(notes, start=42):
        ws.cell(i, 1, line)
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=8)

    for col, w in zip("ABCDEFGH", [36, 14, 55, 16, 16, 14, 12, 26]):
        ws.column_dimensions[col].width = w

    # ---------- Hydrogramme ----------
    wh = wb.create_sheet("Hydrogramme")
    wh["A1"] = "Hydrogramme d'entrée (point Ø1500 / entrée rétention)"
    wh["A1"].font = Font(bold=True, size=12, color="0F5C5C")
    headers = ["t (min)", "Qin (m³/s)", "Remarque"]
    for j, h in enumerate(headers, start=1):
        wh.cell(3, j, h)
    style_header(wh, 3, 3)
    for i, (t, q) in enumerate(HYDRO, start=4):
        wh.cell(i, 1, t).border = THIN
        wh.cell(i, 1).number_format = "0.0"
        wh.cell(i, 2, q).border = THIN
        wh.cell(i, 2).number_format = "0.000"
        note = ""
        if abs(q - 9.6) < 1e-6:
            note = "Pointe"
            wh.cell(i, 2).fill = ORANGE
        elif abs(q - QCAP_900) < 1e-6:
            note = "Qin = Qcap (croisement)"
            wh.cell(i, 2).fill = BLUE
        wh.cell(i, 3, note).border = THIN
    last_h = 3 + len(HYDRO)
    sum_row = last_h + 2
    wh.cell(sum_row, 1, "Qpointe (m³/s)")
    wh.cell(sum_row, 2, f"=MAX(B4:B{last_h})")
    wh.cell(sum_row, 2).fill = GREEN
    wh.cell(sum_row + 1, 1, "t pointe (min)")
    wh.cell(sum_row + 1, 2, 22.5)
    wh.cell(sum_row + 1, 2).number_format = "0.0"

    chart = LineChart()
    chart.title = "Hydrogramme Qin"
    chart.style = 10
    chart.y_axis.title = "Q (m³/s)"
    chart.x_axis.title = "t (min)"
    chart.add_data(Reference(wh, min_col=2, min_row=3, max_row=last_h), titles_from_data=True)
    chart.set_categories(Reference(wh, min_col=1, min_row=4, max_row=last_h))
    chart.shape = 4
    wh.add_chart(chart, "E3")
    for col, w in zip("ABC", [12, 14, 16]):
        wh.column_dimensions[col].width = w

    # ---------- Calcul_A : constant Qout ----------
    wa = wb.create_sheet("Calcul_A")
    wa["A1"] = "Méthode A — Qout constant = capacité Ø900"
    wa["A1"].font = Font(bold=True, size=12, color="0F5C5C")
    wa["A2"] = "Qout (m³/s)"
    wa["B2"] = "=Parametres!B22"
    wa["B2"].fill = YELLOW
    wa["B2"].number_format = "0.00"
    wa["A3"] = "Vmax (m³)"
    n = len(HYDRO)
    first = 8
    last = first + n - 1
    wa["B3"] = f"=MAX(H{first}:H{last})"
    wa["B3"].fill = GREEN
    wa["B3"].number_format = "0.0"
    wa["C3"] = "Maximum du volume stocké cumulé"

    cols = [
        "t (min)",
        "Qin (m³/s)",
        "Qout (m³/s)",
        "Qin−Qout (m³/s)",
        "Δt (s)",
        "ΔVin (m³)",
        "ΔVout (m³)",
        "V stocké (m³)",
        "Stockage?",
    ]
    for j, h in enumerate(cols, start=1):
        wa.cell(7, j, h)
    style_header(wa, 7, len(cols))

    for i, (t, q) in enumerate(HYDRO):
        r = first + i
        # t, Qin from hydro sheet
        wa.cell(r, 1, f"=Hydrogramme!A{4+i}").border = THIN
        wa.cell(r, 1).number_format = "0.0"
        wa.cell(r, 2, f"=Hydrogramme!B{4+i}").border = THIN
        wa.cell(r, 2).number_format = "0.000"
        wa.cell(r, 3, "=$B$2").border = THIN
        wa.cell(r, 3).number_format = "0.00"
        # Signed excess: positive = filling, negative = draining (V floored at 0)
        wa.cell(r, 4, f"=B{r}-C{r}").border = THIN
        wa.cell(r, 4).number_format = "0.000"
        if i == 0:
            wa.cell(r, 5, 0).border = THIN
            wa.cell(r, 6, 0).border = THIN
            wa.cell(r, 7, 0).border = THIN
            wa.cell(r, 8, 0).border = THIN
        else:
            prev = r - 1
            wa.cell(r, 5, f"=(A{r}-A{prev})*60").border = THIN
            wa.cell(r, 5).number_format = "0.0"
            wa.cell(r, 6, f"=0.5*(B{prev}+B{r})*E{r}").border = THIN
            wa.cell(r, 6).number_format = "0.0"
            # Outflow volume consistent with capacity when storing; when empty, limited by Qin
            wa.cell(r, 7, f"=F{r}-(H{r}-H{prev})").border = THIN
            wa.cell(r, 7).number_format = "0.0"
            # Level-pool with constant Qout: V = max(0, V + trap(Qin−Qout)·Δt)
            wa.cell(r, 8, f"=MAX(0,H{prev}+0.5*(D{prev}+D{r})*E{r})").border = THIN
            wa.cell(r, 8).number_format = "0.0"
        wa.cell(r, 8).fill = GREEN
        wa.cell(r, 9, f'=IF(H{r}>0.5,"oui","non")').border = THIN

    wa["A5"] = "t début stockage approx (min)"
    wa["B5"] = "croisement Qin=Qcap dans Hydrogramme (~8.6 min si Q=1.668)"
    wa["A6"] = "t fin accumulation nette (min)"
    wa["B6"] = "croisement descendant (~45.4 min si Q=1.668)"

    # Chart storage
    ch2 = LineChart()
    ch2.title = "Volume stocké cumulé (Méthode A)"
    ch2.y_axis.title = "V (m³)"
    ch2.x_axis.title = "t (min)"
    ch2.add_data(Reference(wa, min_col=8, min_row=7, max_row=last), titles_from_data=True)
    ch2.set_categories(Reference(wa, min_col=1, min_row=first, max_row=last))
    wa.add_chart(ch2, "K7")

    ch3 = LineChart()
    ch3.title = "Qin vs Qout"
    ch3.y_axis.title = "Q (m³/s)"
    ch3.add_data(Reference(wa, min_col=2, min_row=7, max_row=last), titles_from_data=True)
    ch3.add_data(Reference(wa, min_col=3, min_row=7, max_row=last), titles_from_data=True)
    ch3.set_categories(Reference(wa, min_col=1, min_row=first, max_row=last))
    wa.add_chart(ch3, "K22")

    for col in range(1, 10):
        wa.column_dimensions[get_column_letter(col)].width = 14 if col > 1 else 12
    wa.column_dimensions["D"].width = 20
    wa.column_dimensions["H"].width = 18
    wa.column_dimensions["I"].width = 14

    # ---------- Courbe Q=f(H) FHWA HDS-5 (Ø900) ----------
    # Reuse same hydraulics module as ponceau_fosse_model (inlet + outlet control)
    import sys

    sys.path.insert(0, str(HERE.parent / "ponceau_fosse_model"))
    from hydraulics import (  # type: ignore
        CulvertDitchParams,
        pipe_Q,
        pipe_inlet_control_Q,
        pipe_outlet_control_Q,
    )

    entrance = "square_edge"  # unknown entrance → conservative Ke=0.5
    culvert = CulvertDitchParams(
        D=D_900_M,
        L=L_900,
        elev_invert=Z_US_900,
        elev_invert_ds=Z_DS_900,
        elev_max=Z_US_900 + 3.0,
        n_pipe=N_MANNING,
        entrance=entrance,  # type: ignore[arg-type]
        TW=0.0,
        b=1.0,
        z=1.0,
    )
    hw_list = [round(i * 0.05, 2) for i in range(0, 61)]  # 0 → 3.00 m
    rating_rows: list[tuple[float, float, float, float, float, str]] = []
    for hw in hw_list:
        if hw <= 0:
            rating_rows.append((0.0, Z_US_900, 0.0, 0.0, 0.0, "—"))
            continue
        qi = pipe_inlet_control_Q(hw, culvert)
        qo = pipe_outlet_control_Q(hw, culvert)
        qg, ctrl = pipe_Q(hw, culvert)
        rating_rows.append((hw, Z_US_900 + hw, qi, qo, qg, ctrl))

    wq = wb.create_sheet("Courbe_QH_900")
    wq["A1"] = "Courbe de capacité Ø900 — Q = f(H) (FHWA HDS-5 inlet + outlet)"
    wq["A1"].font = Font(bold=True, size=12, color="0F5C5C")
    wq["A2"] = (
        f"Hypothèses: n={N_MANNING}, entrée={entrance} (Ke={culvert.Ke}, inconnu → conservateur), "
        f"TW=0 (exutoire libre), L={L_900} m, S0={S0_900:.5f}. "
        "Q_gouvernant = min(Qinlet, Qoutlet). Sur ce tuyau long, le contrôle outlet domine."
    )
    wq.merge_cells("A2:G3")
    wq["A2"].alignment = Alignment(wrap_text=True, vertical="top")

    wq["A5"] = "Q_plein Manning (réf.)"
    wq["B5"] = Q_FULL_900
    wq["B5"].number_format = "0.000"
    wq["C5"] = "m³/s — section pleine, pente S0 seulement"
    wq["A6"] = "Entrée (hypothèse)"
    wq["B6"] = entrance
    wq["B6"].fill = YELLOW
    wq["C6"] = "square_edge | beveled | groove_headwall — régénérer le xlsx si changé"
    wq["A7"] = "TW (m au-dessus radier aval)"
    wq["B7"] = 0.0
    wq["B7"].fill = YELLOW
    wq["C7"] = "0 = libre; si TW connu, régénérer avec build_retention_excel.py"

    hdr_q = ["HW (m)", "WSE (m)", "Qinlet", "Qoutlet", "Q gouvernant", "Contrôle"]
    for j, h in enumerate(hdr_q, start=1):
        wq.cell(9, j, h)
    style_header(wq, 9, 6)
    first_q = 10
    for i, (hw, wse, qi, qo, qg, ctrl) in enumerate(rating_rows):
        r = first_q + i
        wq.cell(r, 1, hw).number_format = "0.00"
        wq.cell(r, 2, wse).number_format = "0.00"
        wq.cell(r, 3, qi).number_format = "0.000"
        wq.cell(r, 4, qo).number_format = "0.000"
        wq.cell(r, 5, qg).number_format = "0.000"
        wq.cell(r, 6, ctrl)
        for c in range(1, 7):
            wq.cell(r, c).border = THIN
        wq.cell(r, 5).fill = GREEN
    last_q = first_q + len(rating_rows) - 1

    ch_q = LineChart()
    ch_q.title = "Q gouvernant vs HW (Ø900)"
    ch_q.y_axis.title = "Q (m³/s)"
    ch_q.x_axis.title = "HW (m)"
    ch_q.add_data(Reference(wq, min_col=5, min_row=9, max_row=last_q), titles_from_data=True)
    ch_q.add_data(Reference(wq, min_col=3, min_row=9, max_row=last_q), titles_from_data=True)
    ch_q.add_data(Reference(wq, min_col=4, min_row=9, max_row=last_q), titles_from_data=True)
    ch_q.set_categories(Reference(wq, min_col=1, min_row=first_q, max_row=last_q))
    wq.add_chart(ch_q, "H5")

    for col, w in zip("ABCDEF", [10, 10, 10, 10, 14, 12]):
        wq.column_dimensions[col].width = w

    # ---------- Calcul_B : routage avec courbe FHWA ----------
    wb2 = wb.create_sheet("Calcul_B")
    wb2["A1"] = "Méthode B — routage avec Qout = f(H) depuis Courbe_QH_900 (FHWA)"
    wb2["A1"].font = Font(bold=True, size=12, color="0F5C5C")
    wb2["A2"] = (
        "H = V / Aire_bassin. Qout interpolé sur Courbe_QH_900!E (Q gouvernant). "
        "Si V≈0, Qout = MIN(Qin, Q_plein). ΔV = (Qin_moy − Qout)·Δt."
    )
    wb2.merge_cells("A2:G3")
    wb2["A2"].alignment = Alignment(wrap_text=True)

    wb2["A5"] = "Aire (m²)"
    wb2["B5"] = "=Parametres!B24"
    wb2["B5"].fill = YELLOW
    wb2["A6"] = "Q_plein réf. (m³/s)"
    wb2["B6"] = "=Parametres!B22"
    wb2["B6"].fill = YELLOW
    r0 = 12
    wb2["A7"] = "Vmax_B (m³)"
    wb2["B7"] = f"=MAX(G{r0}:G{r0+n-1})"
    wb2["B7"].fill = GREEN
    wb2["B7"].number_format = "0.0"
    wb2["A8"] = "Hmax_B (m)"
    wb2["B8"] = "=IF(B5>0,B7/B5,0)"
    wb2["B8"].fill = BLUE
    wb2["B8"].number_format = "0.00"

    # Named ranges for interpolation (sheet-local absolute refs)
    hw_rng = f"Courbe_QH_900!$A${first_q}:$A${last_q}"
    q_rng = f"Courbe_QH_900!$E${first_q}:$E${last_q}"

    def q_from_h_formula(h_cell: str) -> str:
        # Linear interp between MATCH(...,1) and next row; clamp at ends
        return (
            f'IF({h_cell}<=0,0,'
            f'IF({h_cell}>=INDEX({hw_rng},ROWS({hw_rng})),INDEX({q_rng},ROWS({q_rng})),'
            f'LET('
            f'i,MATCH({h_cell},{hw_rng},1),'
            f'h1,INDEX({hw_rng},i),h2,INDEX({hw_rng},i+1),'
            f'q1,INDEX({q_rng},i),q2,INDEX({q_rng},i+1),'
            f'q1+({h_cell}-h1)/(h2-h1)*(q2-q1))))'
        )

    # LibreOffice-friendly version without LET (nested INDEX/MATCH)
    def q_from_h_classic(h_cell: str) -> str:
        m = f"MATCH({h_cell},{hw_rng},1)"
        return (
            f'IF({h_cell}<=0,0,'
            f'IF({h_cell}>=INDEX({hw_rng},ROWS({hw_rng})),INDEX({q_rng},ROWS({q_rng})),'
            f'INDEX({q_rng},{m})+'
            f'({h_cell}-INDEX({hw_rng},{m}))/'
            f'(INDEX({hw_rng},{m}+1)-INDEX({hw_rng},{m}))*'
            f'(INDEX({q_rng},{m}+1)-INDEX({q_rng},{m}))))'
        )

    hdr_b = ["t (min)", "Qin", "H (m)", "Q_courbe", "Qout", "Δt (s)", "V (m³)"]
    for j, h in enumerate(hdr_b, start=1):
        wb2.cell(11, j, h)
    style_header(wb2, 11, 7)

    for i in range(n):
        r = r0 + i
        wb2.cell(r, 1, f"=Hydrogramme!A{4+i}")
        wb2.cell(r, 2, f"=Hydrogramme!B{4+i}")
        if i == 0:
            wb2.cell(r, 3, 0)
            wb2.cell(r, 4, 0)
            wb2.cell(r, 5, f"=MIN(B{r},$B$6)")
            wb2.cell(r, 6, 0)
            wb2.cell(r, 7, 0)
        else:
            p = r - 1
            wb2.cell(r, 3, f"=IF($B$5>0,G{p}/$B$5,0)")
            wb2.cell(r, 4, f"={q_from_h_classic(f'C{r}')}")
            wb2.cell(r, 5, f"=IF(G{p}<=0.01,MIN(0.5*(B{p}+B{r}),$B$6),D{r})")
            wb2.cell(r, 6, f"=(A{r}-A{p})*60")
            wb2.cell(r, 7, f"=MAX(0,G{p}+(0.5*(B{p}+B{r})-E{r})*F{r})")
        for c in range(1, 8):
            wb2.cell(r, c).border = THIN
            if c in (2, 3, 4, 5):
                wb2.cell(r, c).number_format = "0.000"
            if c in (6, 7):
                wb2.cell(r, c).number_format = "0.0"
        wb2.cell(r, 7).fill = GREEN

    for col, w in zip("ABCDEFG", [10, 10, 10, 12, 10, 10, 12]):
        wb2.column_dimensions[col].width = w

    # ---------- Fichiers liés ----------
    wf = wb.create_sheet("Fichiers_lies")
    wf["A1"] = "Où sont les fichiers HY-8 / capacité ?"
    wf["A1"].font = Font(bold=True, size=13, color="0F5C5C")
    files_txt = [
        "",
        "Ce classeur (rétention Ø900) :",
        "  engineering/drainage-design/retention_volume/Volume_Retention_Ponceau_900.xlsx",
        "",
        "Courbe Q=f(H) FHWA pour CE Ø900 : feuille Courbe_QH_900 (ci-dessus).",
        "",
        "Ancien Excel « type HY-8 / inlet-outlet » (autre projet / Ste-Thérèse) :",
        "  engineering/drainage-design/Ponceau_Capacite_Inlet_Outlet.xlsx",
        "  Branche git: cursor/culvert-excel-capacity-model-d87f",
        "",
        "Modèle composé ponceau Ø1050 + fossé (FHWA + remblai) — PAS le Ø900 de rétention :",
        "  engineering/drainage-design/ponceau_fosse_model/Ponceau_Fosse_Modele_Hydraulique.xlsx",
        "  Branche git: cursor/ponceau-fosse-hydraulic-model-d87f",
        "",
        "HEC-RAS : pas de projet .ras/.hdf pour ce Ø900. Un DEM Ste-Thérèse est dans",
        "  engineering/hec-ras-data/ (données topo seulement).",
        "",
        "Donc: on n'avait pas perdu un HY-8 pour ce bassin Ø900 — on avait des outils",
        "similaires pour d'autres sites. La courbe de CE classeur remplace ça pour le Ø900.",
    ]
    for i, line in enumerate(files_txt, start=2):
        wf.cell(i, 1, line)
    wf.column_dimensions["A"].width = 100

    # ---------- Methode ----------
    wm = wb.create_sheet("Methode")
    wm["A1"] = "Méthode — volume de rétention"
    wm["A1"].font = Font(bold=True, size=13)
    lines = [
        "",
        "1. Hydrogramme Qin(t) à l'entrée de la zone de rétention.",
        "2. Débit sortant contrôlé par le Ø900 (en série avant le Ø1200).",
        "3. Capacité Ø900 — deux niveaux:",
        "   A) Q_plein Manning (n=0.013) ≈ 1.668 m³/s → Méthode A (constant).",
        "   B) Courbe Q=f(H) FHWA HDS-5 (inlet + outlet) → feuille Courbe_QH_900 + Calcul_B.",
        "4. Méthode A: Qout = constante = Q_plein (jaune Parametres). Bon pour 1re estimation.",
        "5. Méthode B: Qout lu sur la courbe selon H=V/Aire (charge amont).",
        "6. Sur ce Ø900 (L/D≈56), le calcul FHWA indique surtout un contrôle outlet.",
        "7. Le Ø1200 est aval: n'augmente pas la sortie de la rétention.",
        "",
        "Fichiers liés / anciens Excel: voir feuille Fichiers_lies.",
    ]
    for i, line in enumerate(lines, start=2):
        wm.cell(i, 1, line)
    wm.column_dimensions["A"].width = 110

    # Link results from Calcul_B on Parametres
    ws["A47"] = "Vmax_B courbe FHWA (m³)"
    ws["B47"] = "=Calcul_B!B7"
    ws["B47"].fill = GREEN
    ws["B47"].number_format = "0"
    ws["C47"] = "Routage avec Q=f(H) — dépend de Aire_bassin (B24)"

    wb.save(OUT)
    return OUT


if __name__ == "__main__":
    path = build()
    print(f"Wrote {path}")
