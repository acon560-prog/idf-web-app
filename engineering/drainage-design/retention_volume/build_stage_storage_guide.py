#!/usr/bin/env python3
"""
Build a teaching Excel: how to link culvert Q = f(WSE) with survey V = f(WSE).

Site numbers from user / boss:
  Ø900 invert US = 34.88 m, D = 0.9 m
  Crown = 35.78 m
  +1.5 m above crown → WSE = 37.28 m → Q = 2.4 m³/s
  "Normal" capacity ≈ 1.0 m³/s
  Ø1500 peak in ≈ 9.6 m³/s; Ø1200 ≈ 4.93 m³/s (downstream, not controlling)
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

HERE = Path(__file__).resolve().parent
OUT = HERE / "Stage_Storage_Discharge_Guide_900.xlsx"

YELLOW = PatternFill("solid", fgColor="FFF59D")
GREEN = PatternFill("solid", fgColor="D1FAE5")
BLUE = PatternFill("solid", fgColor="DBEAFE")
ORANGE = PatternFill("solid", fgColor="FCE4D6")
HEADER_FILL = PatternFill("solid", fgColor="0F5C5C")
HEADER = Font(bold=True, color="FFFFFF")
TITLE = Font(bold=True, size=14, color="0F5C5C")
BOLD = Font(bold=True)
THIN = Border(
    left=Side(style="thin", color="CBD5E1"),
    right=Side(style="thin", color="CBD5E1"),
    top=Side(style="thin", color="CBD5E1"),
    bottom=Side(style="thin", color="CBD5E1"),
)

# Example inundation areas (m²) — placeholders until survey filled in
EXAMPLE_AREAS = {
    34.88: 0,
    35.50: 80,
    36.00: 180,
    36.50: 320,
    37.00: 480,
    37.28: 600,  # REPLACE with boss delineation area at +1.5 m above crown
    37.50: 700,
    39.50: 1032,  # user's polygon — NOT the design link level
}


def style_header(ws, row: int, ncols: int) -> None:
    for c in range(1, ncols + 1):
        cell = ws.cell(row, c)
        cell.font = HEADER
        cell.fill = HEADER_FILL
        cell.border = THIN
        cell.alignment = Alignment(wrap_text=True, vertical="center")


def wrap(ws, cell: str, text: str, merge: str | None = None) -> None:
    ws[cell] = text
    ws[cell].alignment = Alignment(wrap_text=True, vertical="top")
    if merge:
        ws.merge_cells(merge)


def build() -> Path:
    wb = Workbook()

    # ========== 1. START_HERE ==========
    ws = wb.active
    ws.title = "1_START_HERE"
    ws["A1"] = "How flowrate links to stage–storage (Ø900 retention)"
    ws["A1"].font = TITLE
    ws.merge_cells("A1:B1")

    intro = [
        "",
        "WHAT YOUR BOSS ASKED",
        "He delineated a flood limit at about the water level where the Ø900 can pass 2.4 m³/s",
        "(water 1.5 m above the pipe crown). He wants you to CONNECT:",
        "  • how much water is stored on the ground at that level (VOLUME from survey), and",
        "  • how much the culvert can discharge at that same level (FLOWRATE).",
        "",
        "THE ONE RULE",
        "Always use the SAME water surface elevation (WSE) for both:",
        "  WSE → Qout from the culvert     (stage–discharge)",
        "  WSE → Area → Volume from topo   (stage–storage)",
        "Same WSE = the link. Do not pick 39.5 m for Q=2.4 if Q=2.4 belongs to 37.28 m.",
        "",
        "YOUR SITE NUMBERS (Ø900)",
        "  Invert upstream          = 34.88 m",
        "  Diameter D               = 0.90 m",
        "  Crown                    = 34.88 + 0.90 = 35.78 m",
        "  +1.5 m above crown       = 35.78 + 1.50 = 37.28 m   ← boss design stage",
        "  Q at 'normal' head       ≈ 1.0 m³/s",
        "  Q at 37.28 m             = 2.4 m³/s",
        "  Ø1500 upstream capacity  ≈ 9.6 m³/s (inflow can be larger than Ø900)",
        "  Ø1200 downstream         ≈ 4.93 m³/s (does NOT increase outlet of the pond)",
        "",
        "WHY NOT 39.5 m OR 40 m?",
        "Those contours are higher than the head that defines Q = 2.4 m³/s.",
        "Use them only if you study a worse flood that actually reaches that high.",
        "For the '+1.5 m above crown' story, the matching level is 37.28 m.",
        "",
        "SHEETS IN THIS FILE (read in order)",
        "  2_PLAIN_WORDS     — analogy tank + drain (no formulas stress)",
        "  3_STAGE_DISCHARGE — water level → flowrate Q of the Ø900",
        "  4_STAGE_STORAGE   — water level → area → volume V (edit yellow areas)",
        "  5_THE_LINK        — put Q and V on the SAME rows (the table your boss wants)",
        "  6_ROUTING_IDEA    — how the hydrograph uses that table over time",
        "  7_CHECKLIST       — what to measure / deliver",
        "",
        "Yellow cells = you may edit. Green = results. Orange = boss design stage.",
    ]
    for i, line in enumerate(intro, start=2):
        ws.cell(i, 1, line)
        if line in {
            "WHAT YOUR BOSS ASKED",
            "THE ONE RULE",
            "YOUR SITE NUMBERS (Ø900)",
            "WHY NOT 39.5 m OR 40 m?",
            "SHEETS IN THIS FILE (read in order)",
        }:
            ws.cell(i, 1).font = BOLD
            ws.cell(i, 1).fill = BLUE
    ws.column_dimensions["A"].width = 100

    # ========== 2. PLAIN WORDS ==========
    w2 = wb.create_sheet("2_PLAIN_WORDS")
    w2["A1"] = "Plain-words picture (read this if formulas feel abstract)"
    w2["A1"].font = TITLE
    plain = [
        "",
        "Imagine a bathtub (the low ground around the Ø900) with a drain (the Ø900 culvert).",
        "",
        "1) STAGE = how high the water is in the tub (WSE in metres).",
        "2) The higher the water, the more pressure on the drain → the drain passes MORE flow.",
        "   That is stage–discharge:  Q = f(WSE).",
        "3) The higher the water, the more of the landscape is flooded → more VOLUME stored.",
        "   That is stage–storage:    V = f(WSE).",
        "",
        "LINK:",
        "When the water sits at ONE chosen stage (example: 37.28 m), BOTH are true at once:",
        "  • the map shows a certain flooded area / volume V(37.28)",
        "  • the culvert can pass Q(37.28) = 2.4 m³/s",
        "You are not choosing volume from 39.5 m and flow from 37.28 m. Same stage for both.",
        "",
        "DURING A STORM:",
        "  Inflow from upstream (through Ø1500) often exceeds what the Ø900 can pass.",
        "  Extra water piles up → WSE rises → V increases → Qout also increases a bit.",
        "  Peak storage is the highest V (and WSE) during the storm.",
        "  Your boss's 37.28 m line is a CHECK: does the storm peak stay at or below that?",
        "",
        "Ø1200 downstream is like a larger pipe AFTER the drain. It cannot pull more water",
        "through the Ø900 than the Ø900 itself will pass.",
    ]
    for i, line in enumerate(plain, start=2):
        w2.cell(i, 1, line)
        if line.startswith("LINK:") or line.startswith("DURING"):
            w2.cell(i, 1).font = BOLD
            w2.cell(i, 1).fill = BLUE
    w2.column_dimensions["A"].width = 100

    # ========== 3. STAGE DISCHARGE ==========
    w3 = wb.create_sheet("3_STAGE_DISCHARGE")
    w3["A1"] = "Stage–discharge: Q = f(WSE) for the Ø900"
    w3["A1"].font = TITLE
    wrap(
        w3,
        "A2",
        "This table answers: IF the water surface is at this elevation, how much can the Ø900 discharge? "
        "Two office points are fixed: ~1.0 m³/s at lower head, and 2.4 m³/s at 37.28 m. "
        "In-between values are linear placeholders until you refine with HY-8/FHWA.",
        "A2:G3",
    )

    w3["A5"] = "Invert US (m)"
    w3["B5"] = 34.88
    w3["B5"].fill = YELLOW
    w3["A6"] = "D (m)"
    w3["B6"] = 0.90
    w3["B6"].fill = YELLOW
    w3["A7"] = "Crown (m) = Invert + D"
    w3["B7"] = "=B5+B6"
    w3["B7"].fill = GREEN
    w3["B7"].number_format = "0.00"
    w3["A8"] = "Head above crown for Q=2.4 (m)"
    w3["B8"] = 1.50
    w3["B8"].fill = YELLOW
    w3["A9"] = "WSE at Q=2.4 (m)"
    w3["B9"] = "=B7+B8"
    w3["B9"].fill = GREEN
    w3["B9"].number_format = "0.00"
    w3["C9"] = "← this must match the topo contour you delineate for the 2.4 m³/s story"

    w3["A11"] = "Q_normal (m³/s)"
    w3["B11"] = 1.0
    w3["B11"].fill = YELLOW
    w3["A12"] = "Q_at_boss_stage (m³/s)"
    w3["B12"] = 2.4
    w3["B12"].fill = YELLOW

    headers = [
        "WSE (m)",
        "Head above invert H (m)",
        "Head above crown (m)",
        "Qout (m³/s)",
        "Comment",
    ]
    for j, h in enumerate(headers, start=1):
        w3.cell(14, j, h)
    style_header(w3, 14, 5)

    stages_q = [34.88, 35.50, 35.78, 36.00, 36.50, 37.00, 37.28, 37.50, 39.50]
    for i, wse in enumerate(stages_q):
        r = 15 + i
        w3.cell(r, 1, wse).number_format = "0.00"
        w3.cell(r, 2, f"=A{r}-$B$5").number_format = "0.00"
        w3.cell(r, 3, f"=A{r}-$B$7").number_format = "0.00"
        if wse <= 34.88 + 1e-9:
            q_formula = "0"
            note = "Dry / no ponding"
        elif wse < 35.78:
            q_formula = f"$B$11*(A{r}-$B$5)/$B$6"
            note = "Filling toward full pipe (simplified)"
        elif abs(wse - 37.28) < 1e-9:
            q_formula = "$B$12"
            note = "BOSS STAGE: +1.5 m above crown → use THIS for the link"
            w3.cell(r, 1).fill = ORANGE
        elif wse < 37.28:
            q_formula = f"$B$11+($B$12-$B$11)*(A{r}-$B$7)/($B$9-$B$7)"
            note = "Between crown and boss stage (linear placeholder)"
        elif abs(wse - 39.5) < 1e-9:
            q_formula = "$B$12+0.3"
            note = "Your 39.5 m contour — NOT the 2.4 m³/s stage (do not mix)"
            w3.cell(r, 1).fill = BLUE
        else:
            q_formula = f"$B$12+($B$12*0.1)*(A{r}-$B$9)/($B$9-$B$7)"
            note = "Above boss stage (illustrative only)"
        w3.cell(r, 4, f"={q_formula}")
        w3.cell(r, 4).number_format = "0.00"
        w3.cell(r, 4).fill = YELLOW
        if abs(wse - 37.28) < 1e-9:
            w3.cell(r, 4).fill = ORANGE
        w3.cell(r, 5, note)
        for c in range(1, 6):
            w3.cell(r, c).border = THIN

    last_q = 14 + len(stages_q)
    w3["A26"] = "How to read one row"
    w3["A26"].font = BOLD
    wrap(
        w3,
        "A27",
        "Example orange row: WSE = 37.28 m. Column D must be 2.4 m³/s. "
        "When you delineate the flood limit on the survey at 37.28 m, that polygon belongs to THIS row — "
        "same stage as this flowrate. That is the link on the discharge side.",
        "A27:E29",
    )

    chart = LineChart()
    chart.title = "Ø900 stage–discharge Q(WSE)"
    chart.y_axis.title = "Qout (m³/s)"
    chart.x_axis.title = "WSE (m)"
    chart.add_data(Reference(w3, min_col=4, min_row=14, max_row=last_q), titles_from_data=True)
    chart.set_categories(Reference(w3, min_col=1, min_row=15, max_row=last_q))
    w3.add_chart(chart, "G5")

    for col, width in zip("ABCDE", [12, 22, 20, 12, 55]):
        w3.column_dimensions[col].width = width

    # ========== 4. STAGE STORAGE ==========
    w4 = wb.create_sheet("4_STAGE_STORAGE")
    w4["A1"] = "Stage–storage: V = f(WSE) from survey areas"
    w4["A1"].font = TITLE
    wrap(
        w4,
        "A2",
        "For EACH water level below, draw the inundation polygon on the topo and type the AREA (m²) in yellow. "
        "Volume between two levels uses the trapezoid rule in plain text: "
        "V_step = (A_i + A_i+1) / 2 * (WSE_i+1 - WSE_i). "
        "Cumulative V = sum of steps from the bottom up.",
        "A2:H4",
    )

    w4["A6"] = "Plain formula for one step"
    w4["A6"].font = BOLD
    w4["A7"] = "V_step (m³) = (Area_low + Area_high) / 2  ×  (WSE_high − WSE_low)"
    w4["A7"].fill = GREEN
    w4.merge_cells("A7:F7")
    w4["A8"] = "V_cumulative at a level = sum of all V_step from invert up to that level"
    w4.merge_cells("A8:F8")

    hdr = [
        "WSE (m)",
        "Area A (m²)  ← EDIT",
        "ΔWSE (m)",
        "V_step (m³)",
        "V_cumulative (m³)",
        "Role",
    ]
    for j, h in enumerate(hdr, start=1):
        w4.cell(10, j, h)
    style_header(w4, 10, 6)

    stages_v = [34.88, 35.50, 36.00, 36.50, 37.00, 37.28, 37.50, 39.50]
    for i, wse in enumerate(stages_v):
        r = 11 + i
        w4.cell(r, 1, wse).number_format = "0.00"
        area = EXAMPLE_AREAS.get(wse, 0)
        w4.cell(r, 2, area).fill = YELLOW
        w4.cell(r, 2).number_format = "0"
        w4.cell(r, 2).border = THIN
        if i == 0:
            w4.cell(r, 3, 0)
            w4.cell(r, 4, 0)
            w4.cell(r, 5, 0)
            w4.cell(r, 6, "Invert — start (area often 0)")
        else:
            prev = r - 1
            w4.cell(r, 3, f"=A{r}-A{prev}").number_format = "0.00"
            w4.cell(r, 4, f"=0.5*(B{prev}+B{r})*C{r}").number_format = "0.0"
            w4.cell(r, 5, f"=E{prev}+D{r}").number_format = "0.0"
            w4.cell(r, 5).fill = GREEN
            if abs(wse - 37.28) < 1e-9:
                w4.cell(r, 6, "BOSS LIMIT — replace Area with YOUR delineation at 37.28 m")
                for c in range(1, 7):
                    if c != 2:
                        w4.cell(r, c).fill = ORANGE
            elif abs(wse - 39.5) < 1e-9:
                w4.cell(r, 6, "Your 1032 m² example — optional higher stage, not the 2.4 link")
            else:
                w4.cell(r, 6, "Intermediate contour — replace with survey area")
        for c in range(1, 7):
            w4.cell(r, c).border = THIN

    last_v = 10 + len(stages_v)
    w4["A21"] = "Worked meaning of V at 37.28 m"
    w4["A21"].font = BOLD
    wrap(
        w4,
        "A22",
        "Cell E for the 37.28 m row = total water volume (m³) stored in the landscape when the free surface "
        "is at 37.28 m (starting from empty at invert). "
        "REPLACE the yellow Area at 37.28 with the area your boss delineated (or that you redraw at 37.28). "
        "Placeholder areas above are FAKE except 39.5→1032 which you measured — they only show the method.",
        "A22:F24",
    )

    w4["A26"] = "V at boss stage (m³)"
    w4["B26"] = "=E16"  # row 16 = 37.28
    w4["B26"].fill = GREEN
    w4["B26"].number_format = "0"
    w4["C26"] = "← storage available when Qout = 2.4 m³/s (same WSE)"

    chart2 = LineChart()
    chart2.title = "Stage–storage V(WSE)"
    chart2.y_axis.title = "V (m³)"
    chart2.x_axis.title = "WSE (m)"
    chart2.add_data(Reference(w4, min_col=5, min_row=10, max_row=last_v), titles_from_data=True)
    chart2.set_categories(Reference(w4, min_col=1, min_row=11, max_row=last_v))
    w4.add_chart(chart2, "H6")

    for col, width in zip("ABCDEF", [12, 18, 12, 14, 18, 60]):
        w4.column_dimensions[col].width = width

    # ========== 5. THE LINK ==========
    w5 = wb.create_sheet("5_THE_LINK")
    w5["A1"] = "THE LINK — same WSE row has both V and Q"
    w5["A1"].font = TITLE
    wrap(
        w5,
        "A2",
        "This sheet merges stage–storage and stage–discharge. "
        "Read across one row: at this water level, the pond holds V cubic metres AND the Ø900 passes Q m³/s. "
        "That pairing is what 'link volume to flowrate' means.",
        "A2:G3",
    )

    link_hdr = ["WSE (m)", "V (m³)", "Qout (m³/s)", "What this row means in one sentence"]
    for j, h in enumerate(link_hdr, start=1):
        w5.cell(5, j, h)
    style_header(w5, 5, 4)

    link_wsEs = [34.88, 35.50, 36.00, 36.50, 37.00, 37.28, 37.50, 39.50]
    q_row_by_wse = {
        34.88: 15,
        35.50: 16,
        36.00: 18,
        36.50: 19,
        37.00: 20,
        37.28: 21,
        37.50: 22,
        39.50: 23,
    }
    v_row_by_wse = {w: 11 + i for i, w in enumerate(stages_v)}

    for i, wse in enumerate(link_wsEs):
        r = 6 + i
        w5.cell(r, 1, wse).number_format = "0.00"
        vr = v_row_by_wse[wse]
        w5.cell(r, 2, f"='4_STAGE_STORAGE'!E{vr}").number_format = "0.0"
        qr = q_row_by_wse[wse]
        w5.cell(r, 3, f"='3_STAGE_DISCHARGE'!D{qr}").number_format = "0.00"
        if abs(wse - 37.28) < 1e-9:
            w5.cell(
                r,
                4,
                "DESIGN LINK: volume stored at boss flood limit AND culvert Q = 2.4 m³/s",
            )
            for c in range(1, 5):
                w5.cell(r, c).fill = ORANGE
        elif abs(wse - 39.5) < 1e-9:
            w5.cell(
                r,
                4,
                "Higher stage you drew — Q here is NOT the official 2.4 definition stage",
            )
        elif wse <= 34.88:
            w5.cell(r, 4, "Empty start")
        else:
            w5.cell(r, 4, "Intermediate: both V and Q belong to this same water level")
        for c in range(1, 5):
            w5.cell(r, c).border = THIN
        if abs(wse - 37.28) >= 1e-9:
            w5.cell(r, 2).fill = GREEN

    w5["A16"] = "Sentence you can tell your boss"
    w5["A16"].font = BOLD
    wrap(
        w5,
        "A17",
        "At water level 37.28 m (invert 34.88 + diameter 0.90 + 1.50 m above crown), "
        "the survey inundation stores V m³ (see orange row, column B after you enter the real area), "
        "and at that same level the Ø900 can discharge 2.4 m³/s. "
        "Upstream may bring up to ~9.6 m³/s, so volume accumulates whenever Qin > Qout; "
        "the 37.28 m contour is the geometric limit we check against that paired Q.",
        "A17:D20",
    )

    w5["A22"] = "V_boss (m³)"
    w5["B22"] = "=B11"  # 37.28 row
    w5["B22"].fill = GREEN
    w5["B22"].number_format = "0"
    w5["A23"] = "Q_boss (m³/s)"
    w5["B23"] = "=C11"
    w5["B23"].fill = GREEN
    w5["B23"].number_format = "0.00"

    ch_v = LineChart()
    ch_v.title = "V vs WSE (from survey)"
    ch_v.y_axis.title = "V (m³)"
    ch_v.add_data(Reference(w5, min_col=2, min_row=5, max_row=13), titles_from_data=True)
    ch_v.set_categories(Reference(w5, min_col=1, min_row=6, max_row=13))
    w5.add_chart(ch_v, "F5")

    ch_q = LineChart()
    ch_q.title = "Q vs WSE (culvert)"
    ch_q.y_axis.title = "Q (m³/s)"
    ch_q.add_data(Reference(w5, min_col=3, min_row=5, max_row=13), titles_from_data=True)
    ch_q.set_categories(Reference(w5, min_col=1, min_row=6, max_row=13))
    w5.add_chart(ch_q, "F20")

    for col, width in zip("ABCD", [12, 12, 14, 75]):
        w5.column_dimensions[col].width = width

    # ========== 6. ROUTING IDEA ==========
    w6 = wb.create_sheet("6_ROUTING_IDEA")
    w6["A1"] = "How a storm uses the link (routing idea)"
    w6["A1"].font = TITLE
    wrap(
        w6,
        "A2",
        "You do not pick one volume and stop. During the storm, WSE moves up and down the link table. "
        "Each time step: look up Qout from current WSE (or from V→WSE), compare to Qin, update volume.",
        "A2:G3",
    )

    steps = [
        "",
        "Step-by-step logic (every time interval Δt):",
        "  1. Know current stored volume V (starts at 0).",
        "  2. From V, find WSE on the stage–storage curve (sheet 4).",
        "  3. From that WSE, find Qout on the stage–discharge curve (sheet 3).",
        "  4. Qin comes from the hydrograph (Ø1500 inlet hydrograph).",
        "  5. Change in storage:  ΔV = (Qin − Qout) × Δt",
        "     with Δt in seconds, Qin/Qout in m³/s → ΔV in m³.",
        "  6. New volume:  V_new = max(0, V_old + ΔV).",
        "  7. Repeat. The maximum V (and its WSE) is what you compare to 37.28 m / V_boss.",
        "",
        "If Qin > Qout → water rises (moving up the link table).",
        "If Qin < Qout → water falls (moving down the link table).",
        "",
        "At the boss stage specifically:",
        "  When WSE = 37.28 m, Qout = 2.4 m³/s and V = V_boss from your polygon.",
        "  If routing shows WSEmax ≤ 37.28, the delineated limit contains the storm.",
        "  If WSEmax > 37.28, the storm needs more room (or more outlet capacity).",
    ]
    for i, line in enumerate(steps, start=5):
        w6.cell(i, 1, line)
        if line.startswith("Step-by-step") or line.startswith("At the boss"):
            w6.cell(i, 1).font = BOLD
            w6.cell(i, 1).fill = BLUE

    w6["A24"] = "Mini numeric example (one step only)"
    w6["A24"].font = BOLD
    w6["A25"] = "Qin (m³/s)"
    w6["B25"] = 6.0
    w6["B25"].fill = YELLOW
    w6["A26"] = "Current WSE (m)"
    w6["B26"] = 37.28
    w6["B26"].fill = YELLOW
    w6["A27"] = "Qout at that WSE (m³/s)"
    w6["B27"] = 2.4
    w6["B27"].fill = YELLOW
    w6["A28"] = "Δt (minutes)"
    w6["B28"] = 5
    w6["B28"].fill = YELLOW
    w6["A29"] = "Δt (seconds)"
    w6["B29"] = "=B28*60"
    w6["B29"].fill = GREEN
    w6["A30"] = "ΔV this step (m³)"
    w6["B30"] = "=(B25-B27)*B29"
    w6["B30"].fill = GREEN
    w6["B30"].number_format = "0.0"
    w6["C30"] = "Positive → pond gains volume in these 5 minutes"

    wrap(
        w6,
        "A32",
        "Interpretation: with Qin=6 and Qout=2.4, each 5 minutes adds (6−2.4)×300 = 1080 m³. "
        "That is why a peak of 9.6 m³/s needs real storage even though the Ø900 can reach 2.4 m³/s at high stage. "
        "Full hydrograph routing is in Volume_Retention_Ponceau_900.xlsx (Calcul_A / Calcul_B); "
        "this guide explains the V↔Q link your boss asked for.",
        "A32:F35",
    )
    w6.column_dimensions["A"].width = 45
    w6.column_dimensions["B"].width = 14
    w6.column_dimensions["C"].width = 50

    # ========== 7. CHECKLIST ==========
    w7 = wb.create_sheet("7_CHECKLIST")
    w7["A1"] = "Checklist — what to do on the survey / deliver"
    w7["A1"].font = TITLE
    checks = [
        "",
        "□ Confirm arithmetic: crown = 34.88+0.90 = 35.78; boss WSE = 35.78+1.50 = 37.28 m",
        "  (If boss maps 37.38, ask which elevation he used — match Q=2.4 to HIS contour.)",
        "□ Delineate inundation at 37.28 m (or boss contour) and enter Area in sheet 4 (orange row).",
        "□ Add 3–5 intermediate contours between invert and 37.28; enter each Area (yellow).",
        "□ Keep 39.5 / 40 m only as optional 'what if higher' — label them clearly.",
        "□ On sheet 5, show boss the orange row: same WSE → V and Q=2.4 together.",
        "□ Route design hydrograph; report WSEmax and Vmax vs 37.28 / V_boss.",
        "□ State that Ø1200 = 4.93 m³/s is downstream and does not raise pond outlet above Ø900.",
        "",
        "Files:",
        "  This guide:  Stage_Storage_Discharge_Guide_900.xlsx",
        "  Full routing: Volume_Retention_Ponceau_900.xlsx (same folder)",
    ]
    for i, line in enumerate(checks, start=2):
        w7.cell(i, 1, line)
    w7.column_dimensions["A"].width = 100

    wb.save(OUT)
    return OUT


if __name__ == "__main__":
    p = build()
    print(f"Wrote {p}")
