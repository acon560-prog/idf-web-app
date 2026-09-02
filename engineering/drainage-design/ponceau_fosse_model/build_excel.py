"""Build Excel workbook for ponceau + fossé compound model.

Almost everything is live Excel formulas:
- Parametres (D, n, L, …) → Courbe_de_tarage Q(HW) and S_geo
- Hydrogramme / Routing / Resultat_Q9 update from that curve

Python (hydraulics.py + this script) remains available to rebuild the workbook
or validate results; day-to-day use at the office needs only Excel.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from excel_rating_formulas import (
    formula_control,
    formula_q_overflow,
    formula_q_pipe,
    formula_s_geo,
)
from hydraulics import (
    CulvertDitchParams,
    chicago_hydrograph,
    default_triangle_hydrograph,
    route_level_pool,
    solve_HW_for_Q,
)

OUT = Path(__file__).resolve().parent / "Ponceau_Fosse_Modele_Hydraulique.xlsx"

HEADER = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="0F5C5C")
YELLOW = PatternFill("solid", fgColor="FFF59D")
WARN = PatternFill("solid", fgColor="FDE68A")
OK = PatternFill("solid", fgColor="D1FAE5")
THIN = Border(
    left=Side(style="thin", color="CBD5E1"),
    right=Side(style="thin", color="CBD5E1"),
    top=Side(style="thin", color="CBD5E1"),
    bottom=Side(style="thin", color="CBD5E1"),
)

N_TIME_ROWS = 120
N_RATING = 41  # HW points from 0 to Hmax inclusive
RATING_START = 4
RATING_END = RATING_START + N_RATING - 1  # 44


def style_header(ws, row: int, cols: int) -> None:
    for c in range(1, cols + 1):
        cell = ws.cell(row, c)
        cell.font = HEADER
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(wrap_text=True, vertical="center")


def autosize(ws, min_w=12, max_w=42) -> None:
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        width = min_w
        for cell in col:
            if cell.value is None:
                continue
            width = max(width, min(max_w, len(str(cell.value)) + 2))
        ws.column_dimensions[letter].width = width


def excel_interp(
    x: str,
    value_col: str,
    rating_end: int = RATING_END,
    x_col: str = "I",
) -> str:
    """Local linear interpolation on Courbe_de_tarage."""
    xs = f"Courbe_de_tarage!${x_col}$4:${x_col}${rating_end}"
    val = f"Courbe_de_tarage!${value_col}$4:${value_col}${rating_end}"
    m = f"MATCH({x},{xs},1)"
    forecast = (
        f"FORECAST({x},"
        f"INDEX({val},{m}):INDEX({val},{m}+1),"
        f"INDEX({xs},{m}):INDEX({xs},{m}+1))"
    )
    return (
        f"IFERROR({forecast},"
        f"IFERROR(INDEX({val},{m}),INDEX({val},1)))"
    )


def _put_param(ws, row: int, section: str, name: str, value, note: str, yellow: bool = True):
    ws.cell(row, 1, section).border = THIN
    ws.cell(row, 2, name).border = THIN
    cell = ws.cell(row, 3, value)
    cell.border = THIN
    if yellow and not (isinstance(value, str) and value.startswith("=")):
        cell.fill = YELLOW
    ws.cell(row, 4, note).border = THIN


def main() -> None:
    p = CulvertDitchParams()
    Q_design = 10.0  # design peak for Chicago example (editable in Excel)
    pond_area = 50.0
    t_rise = 40.0
    t_fall = 80.0
    dt_min = 5.0
    # Chicago defaults for this culvert/ditch site
    Td_chicago = 60.0  # min — see Formules / Guide
    r_peak = 0.375
    idf_a, idf_b, idf_c = 1000.0, 5.0, 0.80
    rating_end = RATING_END

    t_list, q_list, _intens = chicago_hydrograph(
        Q_peak=Q_design,
        Td_min=Td_chicago,
        r=r_peak,
        a=idf_a,
        b=idf_b,
        c=idf_c,
        dt_min=dt_min,
    )
    route = route_level_pool(t_list, q_list, p, pond_area_m2=pond_area, HW0=0.0)
    peak = max(route, key=lambda s: s.WSE)

    wb = Workbook()

    # --- Parametres (stable cell map used by all live formulas) ---
    ws = wb.active
    ws.title = "Parametres"
    ws["A1"] = "Modèle composé — Ponceau + fossé (remblai) — TOUT LIVE DANS EXCEL"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:D1")
    ws["A2"] = (
        "Cellules jaunes = entrées. Changez D, n, L, cotes… → Courbe_de_tarage + résultats se mettent à jour. "
        "Python (build_excel.py) reste disponible pour régénérer le classeur si besoin."
    )
    ws["A2"].font = Font(italic=True, color="64748B")
    ws.merge_cells("A2:D2")

    for j, h in enumerate(("Section", "Paramètre", "Valeur", "Unité / note"), start=1):
        ws.cell(4, j, h)
    style_header(ws, 4, 4)

    _put_param(ws, 5, "Géométrie", "Diamètre ponceau D", p.D, "m — changez ici sans Python")
    _put_param(ws, 6, "Géométrie", "Longueur L", p.L, "m")
    _put_param(ws, 7, "Géométrie", "Largeur fond fossé b", p.b, "m")
    _put_param(ws, 8, "Géométrie", "Talus z", p.z, "H:V")
    _put_param(ws, 9, "Cote", "Invert amont", p.elev_invert, "m")
    _put_param(ws, 10, "Cote", "Invert aval", p.elev_invert_ds, "m")
    _put_param(ws, 11, "Cote", "Max hauteur d'eau", p.elev_max, "m")
    _put_param(ws, 12, "Cote", "Clé du tuyau (crown)", "=C9+C5", "formule = invert + D", yellow=False)
    _put_param(
        ws, 13, "Hydraulique", "Pente S0", "=MAX(0.000001,(C9-C10)/C6)", "formule = Δz/L", yellow=False
    )
    _put_param(ws, 14, "Hydraulique", "n ponceau", p.n_pipe, "Manning")
    _put_param(ws, 15, "Hydraulique", "Entrée", p.entrance, "beveled | square_edge | groove_headwall")
    _put_param(ws, 16, "Hydraulique", "Tailwater TW", p.TW, "m au-dessus invert aval")
    _put_param(ws, 17, "Débordement", "Mode", p.overflow_mode, "surface | porous | both")
    _put_param(
        ws, 18, "Débordement", "Début débordement y0", "=C5", "m au-dessus invert (défaut = D)", yellow=False
    )
    _put_param(ws, 19, "Débordement", "n surface (rocher)", p.n_surface, "si mode surface/both")
    _put_param(ws, 20, "Débordement", "d50 remblai", p.d50, "m (si porous/both)")
    _put_param(ws, 21, "Débordement", "Porosité", p.n_porosity, "-")
    _put_param(ws, 22, "Débordement", "Cd_rock", p.Cd_rock, "calage poreux")
    _put_param(ws, 23, "Débit", "Q design (permanent)", "=Resultat_Q9!B4", "lié à Resultat_Q9", yellow=False)

    # HDS-5 coeffs auto from entrance (VLOOKUP table)
    ws["A24"] = "Coeffs HDS-5 Form-2 (auto selon Entrée C15)"
    ws["A24"].font = Font(bold=True, color="0F5C5C")
    ws["E4"] = "entrée"
    ws["F4"] = "K"
    ws["G4"] = "M"
    ws["H4"] = "c"
    ws["I4"] = "Y"
    ws["J4"] = "Ke"
    for col in range(5, 11):
        ws.cell(4, col).font = HEADER
        ws.cell(4, col).fill = HEADER_FILL
    presets = [
        ("beveled", 0.0018, 2.5, 0.0300, 0.74, 0.2),
        ("square_edge", 0.0098, 2.0, 0.0398, 0.67, 0.5),
        ("groove_headwall", 0.0078, 2.0, 0.0292, 0.74, 0.2),
    ]
    for i, rowv in enumerate(presets, start=5):
        for j, v in enumerate(rowv, start=5):
            ws.cell(i, j, v)

    _put_param(ws, 25, "HDS-5", "K", '=IFERROR(VLOOKUP(C15,$E$5:$J$7,2,FALSE),0.0018)', "auto", yellow=False)
    _put_param(ws, 26, "HDS-5", "M", '=IFERROR(VLOOKUP(C15,$E$5:$J$7,3,FALSE),2.5)', "auto", yellow=False)
    _put_param(ws, 27, "HDS-5", "c", '=IFERROR(VLOOKUP(C15,$E$5:$J$7,4,FALSE),0.03)', "auto", yellow=False)
    _put_param(ws, 28, "HDS-5", "Y", '=IFERROR(VLOOKUP(C15,$E$5:$J$7,5,FALSE),0.74)', "auto", yellow=False)
    _put_param(ws, 29, "HDS-5", "Ke", '=IFERROR(VLOOKUP(C15,$E$5:$J$7,6,FALSE),0.2)', "auto", yellow=False)

    ws["A31"] = (
        "Sans Python: éditez les jaunes → Courbe_de_tarage / Resultat_Q9 / Routing_stockage (recalcul auto). "
        "Avec Python (optionnel): python build_excel.py régénère ce fichier depuis hydraulics.py."
    )
    ws["A31"].font = Font(italic=True, size=10, color="64748B")
    ws.merge_cells("A31:D32")
    autosize(ws)

    # --- Formules_explications ---
    wsF = wb.create_sheet("Formules_explications", 1)
    wsF["A1"] = "Comment les valeurs sont calculées (formules + méthode)"
    wsF["A1"].font = Font(bold=True, size=14)
    wsF.column_dimensions["A"].width = 28
    wsF.column_dimensions["B"].width = 95

    lines = [
        ("Important", "Presque TOUT est en formules Excel. Changez D/n/L/cotes sur Parametres → la courbe se recalcule. Python reste pour régénérer/valider."),
        ("", ""),
        ("0. Sans Python (bureau)", ""),
        ("Géométrie", "Parametres cellules jaunes C5–C22"),
        ("Q permanent", "Resultat_Q9!B4"),
        ("Hydrogramme", "Hydrogramme_entree: B3=Chicago|Triangle ; B5=Q pointe ; B9=Td ; B10=r ; B11–B13=IDF"),
        ("WSE max orage", "Routing_stockage → WSE max"),
        ("", ""),
        ("Chicago (détail)", ""),
        ("IDF", "ī(td)=a/(td+b)^c"),
        ("i(t) avant pic", "i=a*[(1-c)*(tb/r)+b]/(tb/r+b)^(c+1)   tb=tp-t   tp=r*Td"),
        ("i(t) après pic", "i=a*[(1-c)*(ta/(1-r))+b]/(ta/(1-r)+b)^(c+1)   ta=t-tp"),
        ("Q(t)", "Q=Qpointe*i(t)/i_max   i_max=a/b^c"),
        ("Td choisi", "60 min pour ce site (ponceau court, Tc~10–20 min)"),
        ("", ""),
        ("1. Cotes", ""),
        ("WSE", "WSE = Parametres!C9 + HW"),
        ("S0", "S0 = (invert_amont - invert_aval) / L  (Parametres!C13)"),
        ("", ""),
        ("2. Courbe de tarage", ""),
        ("Création", "Pour chaque HW: Q_total = Q_pipe + Q_overflow — formules Excel (plus de valeurs figées Python)"),
        ("Q_pipe", "MIN(entrée HDS-5, sortie Manning/énergie)"),
        ("Q_overflow", "Manning surface (défaut) ou Wilkins selon Parametres!C17"),
        ("Ind", "Ind = 2*S/dt + Q_out  (lookup routing)"),
        ("", ""),
        ("3. RHS / 2S/dt−Q", ""),
        ("RHS", "RHS = (2S/dt - Q)_prev + Qin_prev + Qin_new"),
        ("2S/dt−Q", "Terme de report Puls (pas un débit)"),
        ("", ""),
        ("4. Python conservé", ""),
        ("hydraulics.py", "Moteur de référence"),
        ("build_excel.py", "Régénère le .xlsx"),
        ("excel_rating_formulas.py", "Texte des formules de courbe"),
        ("run_model.py", "Test terminal"),
    ]
    wsF["A3"] = "Sujet"
    wsF["B3"] = "Formule / explication"
    style_header(wsF, 3, 2)
    for i, (a, b) in enumerate(lines, start=4):
        wsF.cell(i, 1, a)
        wsF.cell(i, 2, b)
        wsF.cell(i, 2).alignment = Alignment(wrap_text=True, vertical="top")
        if a and not b:
            wsF.cell(i, 1).font = Font(bold=True, color="0F5C5C")
        wsF.row_dimensions[i].height = 32 if len(b) > 80 else 18

    # --- Guide sans Python ---
    wsG = wb.create_sheet("Guide_Sans_Python", 2)
    wsG["A1"] = "Guide bureau — Excel seulement (Python optionnel)"
    wsG["A1"].font = Font(bold=True, size=14)
    guide = [
        "",
        "1) Changer la géométrie (D, n, L, inverts…)",
        "   → Parametres, cellules JAUNES → courbe + résultats se mettent à jour",
        "",
        "2) Débit permanent (ex. 9 → 12)",
        "   → Resultat_Q9!B4 jaune → lire WSE",
        "",
        "3) Orage Chicago (défaut) ou Triangle",
        "   → Hydrogramme_entree: B3=Chicago, B5=Q pointe (ex. 10), B9=Td=60 min",
        "   → IDF a,b,c en B11–B13 (caler sur votre station)",
        "   → Routing_stockage « WSE max »",
        "",
        "Chicago en bref: hyétogramme Keifer–Chu → Q(t)=Qpointe*i(t)/i_max",
        "Td=60 min choisi pour ce culvert/fossé (voir Hydrogramme_notes).",
        "",
        "4) Mode débordement",
        "   → Parametres C17 = surface | porous | both",
        "",
        "Python N'EST PAS requis pour ces changements.",
        "Gardez hydraulics.py / build_excel.py pour régénérer le classeur plus tard si besoin.",
    ]
    for i, line in enumerate(guide, start=2):
        wsG.cell(i, 1, line)
    wsG.column_dimensions["A"].width = 100

    # --- Resultat_Q9 LIVE ---
    ws2 = wb.create_sheet("Resultat_Q9")
    ws2["A1"] = "Résultat régime permanent — LIVE (B4 = Q design)"
    ws2["A1"].font = Font(bold=True, size=14)
    ws2["A2"] = "Lit Courbe_de_tarage (calculée depuis Parametres). Changez B4 et/ou D,n,L."
    ws2["A2"].font = Font(italic=True, color="64748B")
    ws2.merge_cells("A2:D2")
    for j, h in enumerate(("Grandeur", "Valeur", "Unité", "Comment obtenu"), start=1):
        ws2.cell(3, j, h)
    style_header(ws2, 3, 4)

    q_cell = "$B$4"
    ws2["A4"] = "Q design"
    ws2["B4"] = Q_design
    ws2["B4"].fill = YELLOW
    ws2["B4"].border = THIN
    ws2["C4"] = "m³/s"
    ws2["D4"] = "ÉDITER ICI"

    zmax_ref = "Parametres!$C$11"
    live_rows = [
        (5, "HW (profondeur amont)", excel_interp(q_cell, "A", x_col="E"), "m", "Q_total → HW"),
        (6, "WSE amont", "=Parametres!$C$9+B5", "m", "invert + HW"),
        (7, "Q_ponceau", excel_interp(q_cell, "C", x_col="E"), "m³/s", "courbe"),
        (8, "Q_débordement", excel_interp(q_cell, "D", x_col="E"), "m³/s", "courbe"),
        (9, "Q_total", "=B7+B8", "m³/s", "≈ B4"),
        (
            10,
            "Contrôle (approx.)",
            f'=IFERROR(INDEX(Courbe_de_tarage!$F$4:$F${rating_end},'
            f'MATCH({q_cell},Courbe_de_tarage!$E$4:$E${rating_end},1)),"-")',
            "-",
            "inlet/outlet",
        ),
        (
            11,
            "Capacité ponceau à la clé",
            excel_interp("Parametres!$C$5", "C", x_col="A"),
            "m³/s",
            "Q_pipe à HW=D",
        ),
        (12, "Débordement actif?", '=IF(B5>Parametres!$C$18,"Oui","Non")', "-", "HW > y0"),
        (13, "Dépasse elev_max?", f'=IF(B6>{zmax_ref},"Oui","Non")', "-", "WSE > max"),
    ]
    for row, label, formula, unit, note in live_rows:
        ws2.cell(row, 1, label).border = THIN
        cell = ws2.cell(row, 2, formula if str(formula).startswith("=") else f"={formula}")
        cell.border = THIN
        cell.fill = OK
        if row in (5, 6, 7, 8, 9, 11):
            cell.number_format = "0.000"
        ws2.cell(row, 3, unit).border = THIN
        ws2.cell(row, 4, note).border = THIN
    ws2.conditional_formatting.add(
        "B13",
        FormulaRule(
            formula=['B13="Oui"'],
            fill=PatternFill(start_color="FDE68A", end_color="FDE68A", fill_type="solid"),
        ),
    )
    ws2["A15"] = "WSE max d'orage → Routing_stockage (pas cette feuille)."
    ws2.merge_cells("A15:D15")
    autosize(ws2)

    # --- Courbe_de_tarage LIVE from Parametres ---
    ws3 = wb.create_sheet("Courbe_de_tarage")
    ws3["A1"] = "Courbe de tarage Q = f(HW) — FORMULES EXCEL depuis Parametres"
    ws3["A1"].font = Font(bold=True, size=14)
    ws3["A2"] = (
        "Chaque ligne recalcule Q_pipe / Q_débord / S_geo quand vous changez D, n, L, cotes… "
        "Plus besoin de Python pour ça."
    )
    ws3["A2"].font = Font(italic=True, color="64748B")
    ws3.merge_cells("A2:I2")
    headers = [
        "HW (m)",
        "WSE",
        "Q_ponceau",
        "Q_débord.",
        "Q_total",
        "Contrôle",
        "S_geo (m³)",
        "S=S_geo+Pond*HW",
        "Ind=2S/dt+Q",
    ]
    for j, h in enumerate(headers, start=1):
        ws3.cell(3, j, h)
    style_header(ws3, 3, 9)

    n_seg = N_RATING - 1
    for i in range(RATING_START, RATING_END + 1):
        k = i - RATING_START
        hw = f"A{i}"
        ws3.cell(
            i, 1, f"=({k}/{n_seg})*(Parametres!$C$11-Parametres!$C$9)"
        ).border = THIN
        ws3.cell(i, 1).number_format = "0.0000"
        ws3.cell(i, 2, f"=Parametres!$C$9+{hw}").border = THIN
        ws3.cell(i, 2).number_format = "0.000"
        ws3.cell(i, 3, f"={formula_q_pipe(hw)}").border = THIN
        ws3.cell(i, 3).number_format = "0.0000"
        ws3.cell(i, 4, f"={formula_q_overflow(hw)}").border = THIN
        ws3.cell(i, 4).number_format = "0.0000"
        ws3.cell(i, 5, f"=C{i}+D{i}").border = THIN
        ws3.cell(i, 5).number_format = "0.0000"
        ws3.cell(i, 6, f"={formula_control(hw)}").border = THIN
        ws3.cell(i, 7, f"={formula_s_geo(hw)}").border = THIN
        ws3.cell(i, 7).number_format = "0.00"
        ws3.cell(i, 8, f"=G{i}+Hydrogramme_entree!$B$4*{hw}").border = THIN
        ws3.cell(i, 8).number_format = "0.00"
        ws3.cell(i, 9, f"=2*H{i}/(Hydrogramme_entree!$B$8*60)+E{i}").border = THIN
        ws3.cell(i, 9).number_format = "0.00"

    ws3.cell(
        rating_end + 2,
        1,
        "Tout est formule. Vérif. optionnelle: python run_model.py",
    )
    ws3.cell(rating_end + 2, 1).font = Font(italic=True, size=10, color="64748B")
    autosize(ws3)

    # --- Hydrogramme (Chicago default + Triangle option) ---
    ws4 = wb.create_sheet("Hydrogramme_entree")
    ws4["A1"] = "Hydrogramme d'entrée — Chicago (défaut) ou Triangle — FORMULES LIVE"
    ws4["A1"].font = Font(bold=True, size=13)
    ws4["A2"] = (
        "Méthode Chicago (Keifer–Chu): hyétogramme IDF → forme de Q(t) calée sur Q pointe. "
        "Durée Td=60 min adaptée à ce ponceau/fossé (Tc site typ. 10–20 min). "
        "Pour Q=10: mettre 10 dans B5. B3=Chicago ou Triangle."
    )
    ws4["A2"].font = Font(italic=True, color="64748B")
    ws4.merge_cells("A2:E2")

    ws4["A3"] = "Méthode"
    c_meth = ws4["B3"]
    c_meth.value = "Chicago"
    c_meth.fill = YELLOW
    c_meth.border = THIN
    ws4["C3"] = "Chicago | Triangle"
    ws4["C3"].font = Font(italic=True, size=9, color="64748B")

    for row, label, val, note in [
        (4, "Pond amont A (m²)", pond_area, "stockage amont (ne pas déplacer: Courbe lit B4)"),
        (5, "Q pointe (m³/s)", Q_design, "ex. 10 — pic de l'hydrogramme"),
        (6, "Montée triangle (min)", t_rise, "utilisé si B3=Triangle"),
        (7, "Descente triangle (min)", t_fall, "utilisé si B3=Triangle"),
        (8, "Pas dt (min)", dt_min, "ne pas déplacer: Courbe/Routing lit B8"),
        (9, "Td Chicago (min)", Td_chicago, "durée orage — défaut 60 min pour ce site"),
        (10, "r (position du pic)", r_peak, "tp=r*Td ; typ. 0.3–0.5 (urbain ~0.375)"),
        (11, "IDF a", idf_a, "ī=a/(t+b)^c  [mm/h, t en min] — caler sur VOTRE IDF"),
        (12, "IDF b", idf_b, "min"),
        (13, "IDF c", idf_c, "-"),
    ]:
        ws4.cell(row, 1, label)
        c = ws4.cell(row, 2, val)
        c.fill = YELLOW
        c.border = THIN
        ws4.cell(row, 3, note).font = Font(italic=True, size=9, color="64748B")

    ws4["A14"] = (
        "Chicago: i(t) (col C) puis Q=Qpointe*i/MAX(i) pour garantir le pic. "
        "Triangle: montée/descente. Voir Formules_explications + Hydrogramme_notes."
    )
    ws4["A14"].font = Font(size=9, color="0F5C5C")
    ws4.merge_cells("A14:E14")

    # B15 = max intensity over the table (for exact Q peak)
    ws4["A15"] = "i_max utilisé (auto)"
    ws4["B15"] = f'=IF(UPPER($B$3)="CHICAGO",MAX(C17:C{17 + N_TIME_ROWS - 1}),1)'
    ws4["B15"].border = THIN
    ws4["C15"] = "MAX des i Chicago — Q=Qpointe*i/i_max"
    ws4["C15"].font = Font(italic=True, size=9, color="64748B")

    def _excel_i(t_cell: str) -> str:
        """Chicago instantaneous intensity at time t_cell."""
        tp = "($B$10*$B$9)"
        tb = f"({tp}-{t_cell})"
        i_b = (
            f"$B$11*((1-$B$13)*(({tb})/$B$10)+$B$12)"
            f"/((({tb})/$B$10)+$B$12)^($B$13+1)"
        )
        ta = f"({t_cell}-{tp})"
        i_a = (
            f"$B$11*((1-$B$13)*(({ta})/(1-$B$10))+$B$12)"
            f"/((({ta})/(1-$B$10))+$B$12)^($B$13+1)"
        )
        return f"IF({t_cell}<={tp},{i_b},{i_a})"

    def _excel_q(t_cell: str, i_cell: str) -> str:
        q_chi = f"IF($B$15<=0,0,$B$5*{i_cell}/$B$15)"
        q_tri = (
            f"IF({t_cell}<=$B$6,$B$5*{t_cell}/$B$6,"
            f"MAX(0,$B$5*(1-({t_cell}-$B$6)/$B$7)))"
        )
        return f'IF(UPPER($B$3)="CHICAGO",{q_chi},{q_tri})'

    def _excel_end() -> str:
        return 'IF(UPPER($B$3)="CHICAGO",$B$9,$B$6+$B$7)'

    ws4["A16"] = "t (min)"
    ws4["B16"] = "Q_in (m³/s)"
    ws4["C16"] = "i Chicago (mm/h)"
    style_header(ws4, 16, 3)

    first_data = 17
    last_data = first_data + N_TIME_ROWS - 1
    end_ref = _excel_end()

    # Build time series: regular dt steps; peak time tp inserted via formula rows
    # Use: t = previous + dt, but also allow hitting tp by using a hybrid:
    # A17=0; next = IF(prev<tp AND prev+dt>tp, tp, IF(prev+dt>end,"",prev+dt))
    tp_ref = "($B$10*$B$9)"

    ws4.cell(first_data, 1, 0).border = THIN
    ws4.cell(
        first_data,
        3,
        f'=IF(OR(A{first_data}="",UPPER($B$3)<>"CHICAGO"),"",{_excel_i(f"A{first_data}")})',
    ).border = THIN
    ws4.cell(
        first_data,
        2,
        f'=IF(A{first_data}="","",{_excel_q(f"A{first_data}", f"C{first_data}")})',
    ).border = THIN
    for col in (2, 3):
        ws4.cell(first_data, col).number_format = "0.0000"
        ws4.cell(first_data, col).fill = PatternFill("solid", fgColor="E0F2FE")

    for r in range(first_data + 1, last_data + 1):
        prev = r - 1
        # Insert peak time tp when the regular step would jump over it
        ws4.cell(
            r,
            1,
            f'=IF(A{prev}="","",'
            f'IF(AND(UPPER($B$3)="CHICAGO",A{prev}<{tp_ref},A{prev}+$B$8>{tp_ref}),{tp_ref},'
            f'IF(A{prev}+$B$8>({end_ref}),"",A{prev}+$B$8)))',
        ).border = THIN
        ws4.cell(
            r,
            3,
            f'=IF(OR(A{r}="",UPPER($B$3)<>"CHICAGO"),"",{_excel_i(f"A{r}")})',
        ).border = THIN
        ws4.cell(
            r,
            2,
            f'=IF(A{r}="","",{_excel_q(f"A{r}", f"C{r}")})',
        ).border = THIN
        for col in (2, 3):
            ws4.cell(r, col).number_format = "0.0000"
            ws4.cell(r, col).fill = PatternFill("solid", fgColor="E0F2FE")
    autosize(ws4)

    # --- Routing ---
    wsR = wb.create_sheet("Routing_stockage")
    wsR["A1"] = "Routing stockage — Puls LIVE (lit Hydrogramme_entree, y compris Chicago)"
    wsR["A1"].font = Font(bold=True, size=13)
    wsR["A2"] = (
        "IMPORTANT: HW = PROFONDEUR (~1.7 m). WSE = ÉLÉVATION = invert + HW (~37 m). "
        "Q_in vient de Hydrogramme_entree (Chicago ou Triangle)."
    )
    wsR["A2"].font = Font(italic=True, color="64748B")
    wsR.merge_cells("A2:K2")
    wsR["A3"] = (
        "Formule: WSE = Parametres!C9 + HW.  2S/dt−Q = report Puls. RHS = cible Ind."
    )
    wsR["A3"].font = Font(italic=True, size=9, color="64748B")
    wsR.merge_cells("A3:K3")

    hdr = [
        "t (min)",
        "Q_in",
        "Q_pipe",
        "Q_overflow",
        "Q_out",
        "HW profondeur (m)",
        "WSE élévation (m)",
        "S stockage (m³)",
        "WSE>max?",
        "RHS (Ind cible)",
        "2S/dt−Q (report Puls)",
    ]
    for j, h in enumerate(hdr, start=1):
        wsR.cell(4, j, h)
    style_header(wsR, 4, 11)

    r0 = 5
    r_last = r0 + N_TIME_ROWS - 1
    dt_ref = "Hydrogramme_entree!$B$8"
    zmax = "Parametres!$C$11"
    zin = "Parametres!$C$9"

    hydro_r = first_data
    wsR.cell(r0, 1, f"=Hydrogramme_entree!A{hydro_r}").border = THIN
    wsR.cell(r0, 2, f"=Hydrogramme_entree!B{hydro_r}").border = THIN
    for col, val in ((3, 0), (4, 0), (5, 0), (6, 0), (8, 0)):
        wsR.cell(r0, col, val).border = THIN
    wsR.cell(r0, 7, f'=IF(A{r0}="","",{zin}+F{r0})').border = THIN
    wsR.cell(r0, 9, f'=IF(G{r0}="","",IF(G{r0}>{zmax},"OUI",""))').border = THIN
    wsR.cell(r0, 10, "").border = THIN
    wsR.cell(r0, 11, f'=IF(A{r0}="","",2*H{r0}/({dt_ref}*60)-E{r0})').border = THIN

    for i in range(1, N_TIME_ROWS):
        r = r0 + i
        prev = r - 1
        hydro_r = first_data + i
        rhs = f"K{prev}+B{prev}+B{r}"

        def gated(formula: str, row: int = r) -> str:
            return f'=IF(A{row}="","",{formula})'

        wsR.cell(r, 1, f"=Hydrogramme_entree!A{hydro_r}").border = THIN
        wsR.cell(r, 2, f"=Hydrogramme_entree!B{hydro_r}").border = THIN
        wsR.cell(r, 3, gated(excel_interp(rhs, "C"))).border = THIN
        wsR.cell(r, 4, gated(excel_interp(rhs, "D"))).border = THIN
        wsR.cell(r, 5, gated(excel_interp(rhs, "E"))).border = THIN
        wsR.cell(r, 6, gated(excel_interp(rhs, "A"))).border = THIN
        wsR.cell(r, 7, f'=IF(A{r}="","",{zin}+F{r})').border = THIN
        wsR.cell(r, 8, gated(excel_interp(rhs, "H"))).border = THIN
        wsR.cell(r, 9, f'=IF(G{r}="","",IF(G{r}>{zmax},"OUI",""))').border = THIN
        wsR.cell(r, 10, gated(rhs)).border = THIN
        wsR.cell(r, 11, f'=IF(A{r}="","",2*H{r}/({dt_ref}*60)-E{r})').border = THIN
        for col in (2, 3, 4, 5, 6, 7, 8, 10, 11):
            wsR.cell(r, col).number_format = "0.0000"

    for col in (2, 3, 4, 5, 6, 7, 8, 11):
        wsR.cell(r0, col).number_format = "0.0000"

    wsR.conditional_formatting.add(
        f"A{r0}:K{r_last}",
        FormulaRule(
            formula=[f'AND($A{r0}<>"",$G{r0}=MAX($G${r0}:$G${r_last}))'],
            fill=PatternFill(start_color="FFF59D", end_color="FFF59D", fill_type="solid"),
        ),
    )

    summary_row = r_last + 2
    wsR.cell(summary_row, 1, "Résumé (live)").font = Font(bold=True)
    wsR.cell(summary_row + 1, 1, "WSE max (élévation)")
    wsR.cell(summary_row + 1, 2, f"=MAX(G{r0}:G{r_last})")
    wsR.cell(summary_row + 1, 2).fill = OK
    wsR.cell(summary_row + 1, 3, "m  ← comparer à 37.4 m")
    wsR.cell(summary_row + 2, 1, "HW max (profondeur)")
    wsR.cell(summary_row + 2, 2, f"=MAX(F{r0}:F{r_last})")
    wsR.cell(summary_row + 2, 3, "m  ← au-dessus de l'invert (≈ WSE−35.36)")
    wsR.cell(summary_row + 3, 1, "S max")
    wsR.cell(summary_row + 3, 2, f"=MAX(H{r0}:H{r_last})")
    wsR.cell(summary_row + 3, 3, "m³")
    wsR.cell(summary_row + 4, 1, "Q_in max")
    wsR.cell(summary_row + 4, 2, f"=MAX(B{r0}:B{r_last})")
    wsR.cell(summary_row + 5, 1, "Réf. Python au build")
    wsR.cell(summary_row + 5, 2, round(peak.WSE, 3))
    wsR.cell(summary_row + 5, 3, "m WSE (validation optionnelle)")
    autosize(wsR)

    # --- Notes ---
    wsN = wb.create_sheet("Hydrogramme_notes")
    wsN["A1"] = "Chicago vs Triangle — et comment c'est construit"
    wsN["A1"].font = Font(bold=True, size=13)
    for i, line in enumerate(
        [
            "",
            "=== MÉTHODE CHICAGO (Keifer–Chu) — défaut du classeur ===",
            "1) IDF: intensité moyenne ī(td) = a / (td + b)^c   (td en minutes, ī en mm/h)",
            "2) Hyétogramme: intensité INSTANTANÉE i(t) avant/après le pic tp = r*Td",
            "   Avant pic (tb = tp − t):  i = a * [(1−c)*(tb/r) + b] / (tb/r + b)^(c+1)",
            "   Après pic (ta = t − tp):  i = a * [(1−c)*(ta/(1−r)) + b] / (ta/(1−r) + b)^(c+1)",
            "3) Pic d'intensité: i_max = a / b^c",
            "4) Hydrogramme calé sur un débit donné:  Q(t) = Q_pointe * i(t) / i_max",
            "   → pour Q_pointe = 10 m³/s, le maximum de Q_in est exactement 10.",
            "",
            "Durée Td = 60 min (choix pour ce ponceau + fossé):",
            "  - temps de parcours dans le tuyau (L=50.9 m) ≈ quelques minutes",
            "  - Tc typique petit bassin urbain/industriel ≈ 10–20 min",
            "  - règle pratique: Td ≥ ~2–3 × Tc → 60 min est adéquat (120 min si bassin plus grand)",
            "",
            "Coeffs a,b,c: remplacer par ceux de VOTRE courbe IDF (période de retour du projet).",
            "La forme de Q(t) dépend surtout de b, c, r, Td ; a s'annule dans i/i_max.",
            "",
            "=== TRIANGLE (option) ===",
            "Mettre B3 = Triangle. Qin = montée/descente linéaires (B6, B7).",
            "",
            "Routing: feuille Routing_stockage lit la colonne Q_in (inchangée).",
        ],
        start=2,
    ):
        wsN.cell(i, 1, line)
    wsN.column_dimensions["A"].width = 110

    # --- Comparaison modes ---
    ws5 = wb.create_sheet("Comparaison_modes")
    ws5["A1"] = "Comparaison modes (instantané Python au build — pédagogique)"
    ws5["A1"].font = Font(bold=True, size=12)
    ws5["A2"] = (
        "Mode du modèle principal: Parametres!C17 (LIVE). "
        "Ce tableau compare les 3 modes à Q fixe au dernier build_excel.py."
    )
    ws5["A2"].font = Font(italic=True, color="64748B")
    ws5.merge_cells("A2:G2")
    ws5["A3"] = "Q comparaison (build)"
    ws5["B3"] = Q_design
    hdr2 = ["Mode", "Signification", "WSE (m)", "Q_pipe", "Q_overflow", "Q_total", "Sous max?"]
    for j, h in enumerate(hdr2, start=1):
        ws5.cell(5, j, h)
    style_header(ws5, 5, 7)
    mode_notes = {
        "porous": "Seulement pores du remblai — faible débit",
        "surface": "Manning fossé — DÉFAUT",
        "both": "Poreux + surface",
    }
    for i, mode in enumerate(("porous", "surface", "both"), start=6):
        pm = CulvertDitchParams(overflow_mode=mode)
        rr = solve_HW_for_Q(Q_design, pm)
        ok = (
            "Oui"
            if rr.WSE <= pm.elev_max + 1e-6 and rr.Q_total >= Q_design - 1e-3
            else "Non / limite"
        )
        vals = [
            mode,
            mode_notes[mode],
            round(rr.WSE, 3),
            round(rr.Q_pipe, 3),
            round(rr.Q_ditch, 3),
            round(rr.Q_total, 3),
            ok,
        ]
        for j, v in enumerate(vals, start=1):
            cell = ws5.cell(i, j, v)
            cell.border = THIN
            if mode == "surface":
                cell.fill = OK
            elif ok.startswith("Non"):
                cell.fill = WARN
    autosize(ws5)

    wb.save(OUT)
    print(f"Wrote {OUT}")
    print(f"Rating rows {RATING_START}..{RATING_END} are live Excel formulas from Parametres")
    print(
        f"Python ref routing peak WSE={peak.WSE:.3f} m at t={peak.t_min:.0f} "
        f"(Qin={peak.Q_in:.2f}, Qout={peak.Q_out:.2f})"
    )


if __name__ == "__main__":
    main()
