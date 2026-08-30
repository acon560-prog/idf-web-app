#!/usr/bin/env python3
"""
Ste-Thérèse ditch — real surveyed sections + longitudinal profile.
Compute Manning normal-depth velocity for Q=7.36 m³/s, n=0.035.

Policy on slopes:
  - Display the real signed S0 from the longitudinal survey (adverse stays negative).
  - Compute Manning velocity only where S0 > 0.
  - Where S0 <= 0 or missing: leave y / WSE / V blank (no |S0|).
"""

from __future__ import annotations

import csv
import math
import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from profile_velocity import SectionGeom, solve_normal_depth

UPLOADS = Path("/home/ubuntu/.cursor/projects/workspace/uploads")
OUT_DIR = Path(__file__).resolve().parent / "st_therese_data"
OUT_XLSX = Path(__file__).resolve().parent / "SteTherese_Fosse_Vitesse_Q736.xlsx"

Q_DESIGN = 7.36
N_MANNING = 0.035

# Map section labels → cumulative station from Longitudinal_5712.txt
SECTION_STATIONS = {
    "1+000": 0.0,
    "2+000": 13.783,
    "3+000": 23.663,
    "4+000": 35.641,
    "5+000": 45.838,
    "6+000": 59.435,
    "7+000": 71.973,
    "8+000": 89.361,
}

HEADER = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="0F5C5C")
YELLOW = PatternFill("solid", fgColor="FFF59D")
OK = PatternFill("solid", fgColor="D1FAE5")
SKIP = PatternFill("solid", fgColor="FCE4D6")  # adverse / flat — no V
WARN = PatternFill("solid", fgColor="FDE68A")
THIN = Border(
    left=Side(style="thin", color="CBD5E1"),
    right=Side(style="thin", color="CBD5E1"),
    top=Side(style="thin", color="CBD5E1"),
    bottom=Side(style="thin", color="CBD5E1"),
)


def style_header(ws, row: int, cols: int) -> None:
    for c in range(1, cols + 1):
        cell = ws.cell(row, c)
        cell.font = HEADER
        cell.fill = HEADER_FILL


def parse_longitudinal(path: Path) -> list[dict]:
    rows = []
    with path.open(encoding="utf-8") as f:
        lines = [ln.strip() for ln in f if ln.strip()]
    # header: Point, Z fond, ΔL, S0, Station cumulées
    for ln in lines[1:]:
        parts = re.split(r"\t+", ln)
        if len(parts) < 5:
            parts = re.split(r"\s{2,}|\t+", ln)
        if len(parts) < 5:
            parts = ln.replace(",", ".").split()
        point = parts[0]
        z = float(parts[1].replace(",", "."))
        s0_raw = parts[3]
        sta = float(parts[4].replace(",", "."))
        if "—" in s0_raw or s0_raw.strip() in ("—", "-", ""):
            s0 = None
        else:
            s0 = float(s0_raw.replace(",", "."))
        try:
            dL = float(parts[2].replace(",", ".")) if "—" not in parts[2] else None
        except ValueError:
            dL = None
        rows.append({"point": point, "z": z, "dL": dL, "S0": s0, "station": sta})
    return rows


def parse_sections(path: Path) -> dict[str, list[tuple[float, float, str]]]:
    """Return {section_id: [(distance, elev, desc), ...]} sorted by distance."""
    text = path.read_text(encoding="utf-8")
    sections: dict[str, list[tuple[float, float, str]]] = {}
    current = None
    for raw in text.splitlines():
        line = raw.strip()
        if not line or line.startswith("Est-ce que"):
            continue
        m = re.match(r"SECTION\s+(\d+\+\d+)", line, re.I)
        if m:
            current = m.group(1)
            sections[current] = []
            continue
        if current is None:
            continue
        if line.lower().startswith("point"):
            continue
        parts = re.split(r"\s+", line)
        if len(parts) < 4:
            continue
        try:
            dist = float(parts[-1].replace(",", "."))
            elev = float(parts[1].replace(",", "."))
        except ValueError:
            continue
        desc = parts[2]
        sections[current].append((dist, elev, desc))
    for k in sections:
        sections[k].sort(key=lambda t: t[0])
    return sections


def s0_at_station(long_rows: list[dict], station: float) -> tuple[float | None, str]:
    """Return real signed S0 at/near station (no abs). None if missing."""
    for r in long_rows:
        if abs(r["station"] - station) < 0.05 and r["S0"] is not None:
            s = r["S0"]
            if s <= 0:
                return s, "S0 ≤ 0 — profondeur normale non applicable"
            return s, "S0 au profil longitudinal"
    candidates = [r for r in long_rows if r["S0"] is not None]
    if not candidates:
        return None, "S0 manquant"
    nearest = min(candidates, key=lambda r: abs(r["station"] - station))
    s = nearest["S0"]
    if s is None or s <= 0:
        return s, f"S0 (station {nearest['station']:.2f} m) ≤ 0 — yn non applicable"
    return s, f"S0 interpolé / station {nearest['station']:.2f} m"


def _seg_area_formula(wse_cell: str, x0: str, z0: str, x1: str, z1: str) -> str:
    """Excel formula: wet area of one polyline segment at WSE (same logic as section_AP)."""
    d0 = f"({wse_cell}-{z0})"
    d1 = f"({wse_cell}-{z1})"
    # intersection parameter t = d0/(d0-d1); xi = x0 + t*(x1-x0)
    t = f"({d0}/({d0}-{d1}))"
    xi = f"({x0}+{t}*({x1}-{x0}))"
    both_dry = f"AND({d0}<=0,{d1}<=0)"
    both_wet = f"AND({d0}>=0,{d1}>=0)"
    a_full = f"0.5*({d0}+{d1})*ABS({x1}-{x0})"
    a_left = f"0.5*{d0}*ABS({xi}-{x0})"
    a_right = f"0.5*{d1}*ABS({x1}-{xi})"
    return (
        f'IF({both_dry},0,'
        f'IF({both_wet},{a_full},'
        f'IF({d0}>0,{a_left},{a_right})))'
    )


def _seg_peri_formula(wse_cell: str, x0: str, z0: str, x1: str, z1: str) -> str:
    """Excel formula: wetted length of one polyline segment at WSE."""
    d0 = f"({wse_cell}-{z0})"
    d1 = f"({wse_cell}-{z1})"
    t = f"({d0}/({d0}-{d1}))"
    xi = f"({x0}+{t}*({x1}-{x0}))"
    both_dry = f"AND({d0}<=0,{d1}<=0)"
    both_wet = f"AND({d0}>=0,{d1}>=0)"
    p_full = f"SQRT(({x1}-{x0})^2+({z1}-{z0})^2)"
    p_left = f"SQRT(({xi}-{x0})^2+({wse_cell}-{z0})^2)"
    p_right = f"SQRT(({x1}-{xi})^2+({wse_cell}-{z1})^2)"
    return (
        f'IF({both_dry},0,'
        f'IF({both_wet},{p_full},'
        f'IF({d0}>0,{p_left},{p_right})))'
    )


def _area_sum_formula(wse_ref: str, first_row: int, n_pts: int, xcol: str = "B", zcol: str = "C") -> str:
    """Sum of segment wet areas at WSE."""
    parts = []
    for i in range(n_pts - 1):
        r = first_row + i
        parts.append(
            f"({_seg_area_formula(wse_ref, f'{xcol}{r}', f'{zcol}{r}', f'{xcol}{r + 1}', f'{zcol}{r + 1}')})"
        )
    return "+".join(parts)


def _peri_sum_formula(wse_ref: str, first_row: int, n_pts: int, xcol: str = "B", zcol: str = "C") -> str:
    parts = []
    for i in range(n_pts - 1):
        r = first_row + i
        parts.append(
            f"({_seg_peri_formula(wse_ref, f'{xcol}{r}', f'{zcol}{r}', f'{xcol}{r + 1}', f'{zcol}{r + 1}')})"
        )
    return "+".join(parts)


def add_area_formula_sheets(
    wb: Workbook,
    sections: dict[str, list[tuple[float, float, str]]],
    results: list,
) -> None:
    """Public section sheets + veryHidden calc sheets (normal depth solver)."""
    N_BISECT = 30
    FMT3 = "0.000"
    FMT5 = "0.00000"

    # Short professional notes (no solver jargon on the face of the workbook)
    wsF = wb.create_sheet("Notes_calcul", 1)
    wsF["A1"] = "Notes de calcul — fossé Ste-Thérèse"
    wsF["A1"].font = Font(bold=True, size=13, color="0F5C5C")
    lines = [
        "",
        "Hypothèses",
        "  • Régime permanent et uniforme (profondeur normale) lorsque la pente longitudinale S0 est positive.",
        "  • Formule de Manning :  Q = (1/n) · A · R^(2/3) · √S0",
        "  • Vitesse :  V = Q / A",
        "  • Géométrie de chaque coupe : polyligne Distance–Élévation du levé (pas un trapèze forfaitaire).",
        "",
        "Aire mouillée A et périmètre P",
        "  Pour une cote d'eau donnée, chaque segment de la coupe est traité comme trapèze (entièrement",
        "  noyé) ou triangle (partiellement noyé). A et P sont les sommes sur tous les segments mouillés.",
        "  R = A / P.",
        "",
        "Cote d'eau (profondeur normale)",
        "  La cote est celle qui vérifie Q_Manning = Q de projet pour la coupe, n et S0 donnés.",
        "  Le calcul est intégré dans le classeur (feuilles techniques masquées).",
        "",
        "Pente adverse (S0 ≤ 0)",
        "  La profondeur normale n'est pas définie ; S0 est reporté tel quel et V n'est pas calculée.",
        "",
        "Entrées (jaune) : Q et n sur la feuille Resultats_Q736 uniquement.",
        "  Modifier Q ou n → toutes les feuilles Section_* et le tableau récapitulatif se mettent à jour.",
    ]
    for i, line in enumerate(lines, start=2):
        wsF.cell(i, 1, line)
    wsF.column_dimensions["A"].width = 100

    s0_by_sid = {sid: S0 for sid, _n, S0, _c, _r in results}

    for sid in sorted(SECTION_STATIONS.keys(), key=lambda x: SECTION_STATIONS[x]):
        pts = sections[sid]
        n_pts = len(pts)
        safe = sid.replace("+", "")
        pub_name = f"Section_{safe}"
        calc_name = f"calc_{safe}"
        s0 = s0_by_sid.get(sid)
        can_normal = s0 is not None and s0 > 0.0
        bed = min(p[1] for p in pts)

        # ----- Public sheet (what the office sees) -----
        ws = wb.create_sheet(pub_name)
        ws["A1"] = f"Section {sid} — vitesse à profondeur normale"
        ws["A1"].font = Font(bold=True, size=12, color="0F5C5C")
        ws.merge_cells("A1:D1")

        ws["A3"] = "Station (m)"
        ws["B3"] = round(SECTION_STATIONS[sid], 3)
        ws["B3"].number_format = FMT3
        ws["A4"] = "S0 (-)"
        ws["B4"] = round(s0, 5) if s0 is not None else None
        if s0 is not None:
            ws["B4"].number_format = FMT5
        if not can_normal:
            ws["B4"].fill = SKIP
            ws["C4"] = "Pente adverse ou nulle — V non calculée"
            ws["C4"].font = Font(italic=True, color="9A3412")

        ws["A5"] = "Q (m³/s)"
        ws["B5"] = "='Resultats_Q736'!B4"
        ws["B5"].border = THIN
        ws["B5"].number_format = FMT3
        ws["C5"] = "lié à Resultats_Q736"
        ws["C5"].font = Font(italic=True, color="64748B", size=9)
        ws["A6"] = "n Manning"
        ws["B6"] = "='Resultats_Q736'!B5"
        ws["B6"].border = THIN
        ws["B6"].number_format = FMT3
        ws["C6"] = "lié à Resultats_Q736"
        ws["C6"].font = Font(italic=True, color="64748B", size=9)

        ws["A8"] = "Résultats"
        ws["A8"].font = Font(bold=True)
        labels = [
            (9, "Cote d'eau WSE (m)"),
            (10, "Profondeur y (m)"),
            (11, "Aire A (m²)"),
            (12, "Périmètre P (m)"),
            (13, "Rayon R = A/P (m)"),
            (14, "Vitesse V (m/s)"),
        ]
        for row, lab in labels:
            ws.cell(row, 1, lab)

        # Geometry reference (survey points only)
        ws["A17"] = "Géométrie du levé"
        ws["A17"].font = Font(bold=True)
        for j, h in enumerate(["#", "Distance (m)", "Élévation (m)", "Desc."], start=1):
            ws.cell(18, j, h)
        style_header(ws, 18, 4)
        for i, (x, z, desc) in enumerate(pts):
            r = 19 + i
            ws.cell(r, 1, i + 1).border = THIN
            c = ws.cell(r, 2, round(x, 3))
            c.border = THIN
            c.number_format = FMT3
            c = ws.cell(r, 3, round(z, 3))
            c.border = THIN
            c.number_format = FMT3
            ws.cell(r, 4, desc).border = THIN

        for col, w in zip("ABCD", [26, 14, 16, 28]):
            ws.column_dimensions[col].width = w

        if not can_normal:
            for row in range(9, 15):
                ws.cell(row, 2, None).fill = SKIP
                ws.cell(row, 2).border = THIN
            continue

        # ----- Hidden calc sheet -----
        wc = wb.create_sheet(calc_name)
        wc.sheet_state = "veryHidden"
        wc["A1"] = f"calc {sid}"
        # Mirror inputs from public sheet
        wc["A3"] = "Q"
        wc["B3"] = f"='{pub_name}'!B5"
        wc["A4"] = "n"
        wc["B4"] = f"='{pub_name}'!B6"
        wc["A5"] = "S0"
        wc["B5"] = f"='{pub_name}'!B4"

        # Points at rows 10+
        wc["A9"] = "x"
        wc["B9"] = "z"
        first = 10
        for i, (x, z, _desc) in enumerate(pts):
            r = first + i
            wc.cell(r, 1, round(x, 4))
            wc.cell(r, 2, round(z, 4))
        last_pt = first + n_pts - 1

        wc["D3"] = "fond"
        wc["E3"] = f"=MIN(B{first}:B{last_pt})"
        wc["D4"] = "zmax"
        wc["E4"] = f"=MAX(B{first}:B{last_pt})"

        # Bisection starting row 10, columns D–J
        # D=lo E=hi F=mid G=A H=P I=Qman
        # Use x in col A, z in col B on calc sheet
        r0 = 10
        # Place bisection to the right so it doesn't overwrite points — use columns D+
        # Actually points are A,B rows 10.. — put bisection starting at row 40
        r0 = last_pt + 3
        wc.cell(r0 - 1, 4, "lo")
        wc.cell(r0 - 1, 5, "hi")
        wc.cell(r0 - 1, 6, "mid")
        wc.cell(r0 - 1, 7, "A")
        wc.cell(r0 - 1, 8, "P")
        wc.cell(r0 - 1, 9, "Qman")

        wc.cell(r0, 4, f"=E3+0.0001")
        wc.cell(r0, 5, f"=E4-0.0001")
        wc.cell(r0, 6, f"=(D{r0}+E{r0})/2")
        wc.cell(r0, 7, f"={_area_sum_formula(f'F{r0}', first, n_pts, 'A', 'B')}")
        wc.cell(r0, 8, f"={_peri_sum_formula(f'F{r0}', first, n_pts, 'A', 'B')}")
        wc.cell(
            r0,
            9,
            f'=IF(OR($B$4<=0,G{r0}<=0,H{r0}<=0,$B$5<=0),0,'
            f'(1/$B$4)*G{r0}*((G{r0}/H{r0})^(2/3))*SQRT($B$5))',
        )
        for k in range(1, N_BISECT):
            r = r0 + k
            prev = r - 1
            wc.cell(r, 4, f'=IF(I{prev}<$B$3,F{prev},D{prev})')
            wc.cell(r, 5, f'=IF(I{prev}<$B$3,E{prev},F{prev})')
            wc.cell(r, 6, f"=(D{r}+E{r})/2")
            wc.cell(r, 7, f"={_area_sum_formula(f'F{r}', first, n_pts, 'A', 'B')}")
            wc.cell(r, 8, f"={_peri_sum_formula(f'F{r}', first, n_pts, 'A', 'B')}")
            wc.cell(
                r,
                9,
                f'=IF(OR($B$4<=0,G{r}<=0,H{r}<=0,$B$5<=0),0,'
                f'(1/$B$4)*G{r}*((G{r}/H{r})^(2/3))*SQRT($B$5))',
            )
        last_bis = r0 + N_BISECT - 1

        # Final outputs on calc sheet
        wc["K3"] = "WSE"
        wc["L3"] = f"=F{last_bis}"
        wc["K4"] = "A"
        wc["L4"] = f"={_area_sum_formula('L3', first, n_pts, 'A', 'B')}"
        wc["K5"] = "P"
        wc["L5"] = f"={_peri_sum_formula('L3', first, n_pts, 'A', 'B')}"
        wc["K6"] = "R"
        wc["L6"] = "=IF(L5>0,L4/L5,0)"
        wc["K7"] = "V"
        wc["L7"] = "=IF(L4>0,B3/L4,0)"
        wc["K8"] = "y"
        wc["L8"] = "=L3-E3"

        # Public results link to hidden calc (3 decimals)
        ws["B9"] = f"='{calc_name}'!L3"
        ws["B10"] = f"='{calc_name}'!L8"
        ws["B11"] = f"='{calc_name}'!L4"
        ws["B12"] = f"='{calc_name}'!L5"
        ws["B13"] = f"='{calc_name}'!L6"
        ws["B14"] = f"='{calc_name}'!L7"
        for row in range(9, 15):
            ws.cell(row, 2).fill = OK
            ws.cell(row, 2).border = THIN
            ws.cell(row, 2).number_format = FMT3


def export_clean_csvs(long_rows: list[dict], sections: dict) -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    with (OUT_DIR / "Longitudinal_st_therese.csv").open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["station_m", "elev_m", "S0_raw", "point"])
        for r in long_rows:
            w.writerow([r["station"], r["z"], r["S0"] if r["S0"] is not None else "", r["point"]])

    for sid, pts in sections.items():
        sta = SECTION_STATIONS[sid]
        name = f"section_{sid.replace('+','')}_sta{sta:.0f}.csv"
        with (OUT_DIR / name).open("w", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            w.writerow(["station_m", "offset_m", "elev_m", "desc"])
            for dist, elev, desc in pts:
                w.writerow([sta, dist, elev, desc])


def main() -> None:
    long_path = UPLOADS / "Longitudinal_5712.txt"
    sec_path = UPLOADS / "sections_stetheres_une_par_une_1c79.txt"
    long_rows = parse_longitudinal(long_path)
    sections = parse_sections(sec_path)
    export_clean_csvs(long_rows, sections)

    results = []
    for sid in sorted(SECTION_STATIONS.keys(), key=lambda x: SECTION_STATIONS[x]):
        sta = SECTION_STATIONS[sid]
        pts = sections[sid]
        offs = [p[0] for p in pts]
        elevs = [p[1] for p in pts]
        sec = SectionGeom(station_m=sta, offsets=offs, elevs=elevs)
        S0_real, snote = s0_at_station(long_rows, sta)
        if S0_real is not None and S0_real > 0.0:
            r = solve_normal_depth(sec, Q_DESIGN, N_MANNING, S0_real)
            results.append((sid, snote, S0_real, True, r))
        else:
            # Keep real S0 (negative or None); do not invent |S0| or a velocity
            bed = min(elevs)
            results.append(
                (
                    sid,
                    snote,
                    S0_real,
                    False,
                    type("R", (), {
                        "station_m": sta,
                        "bed_elev": bed,
                        "S0": S0_real,
                        "y_m": float("nan"),
                        "WSE": float("nan"),
                        "A": float("nan"),
                        "P": float("nan"),
                        "R": float("nan"),
                        "V": float("nan"),
                        "Fr": float("nan"),
                        "ok": False,
                        "note": "pente adverse/nulle — Manning non appliqué",
                    })(),
                )
            )

    # Print summary
    print(f"Q={Q_DESIGN} m³/s  n={N_MANNING}")
    print(f"{'Sec':>8} {'Sta':>8} {'S0_reel':>10} {'y':>7} {'WSE':>8} {'A':>7} {'V':>7} {'Fr':>6}  status")
    for sid, snote, S0_real, computed, r in results:
        s0_str = f"{S0_real:10.5f}" if S0_real is not None else f"{'—':>10}"
        if computed and not math.isnan(r.V):
            print(
                f"{sid:>8} {r.station_m:8.2f} {s0_str} {r.y_m:7.3f} {r.WSE:8.3f} "
                f"{r.A:7.3f} {r.V:7.3f} {r.Fr:6.2f}  {r.note}"
            )
        else:
            print(
                f"{sid:>8} {r.station_m:8.2f} {s0_str} {'—':>7} {'—':>8} "
                f"{'—':>7} {'—':>7} {'—':>6}  {snote}"
            )

    # Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Resultats_Q736"
    ws["A1"] = "Fossé Ste-Thérèse — vitesse aux sections réelles (survey)"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = (
        f"Valeurs par défaut Q = {Q_DESIGN} m³/s ; n = {N_MANNING}. "
        "Modifier Q ou n (jaune) → le tableau et les Section_* se recalculent (pas besoin de Python)."
    )
    ws["A2"].font = Font(italic=True, color="64748B")
    ws.merge_cells("A2:K2")

    ws["A4"] = "Q (m³/s)"
    ws["B4"] = Q_DESIGN
    ws["B4"].fill = YELLOW
    ws["B4"].border = THIN
    ws["B4"].number_format = "0.000"
    ws["A5"] = "n Manning"
    ws["B5"] = N_MANNING
    ws["B5"].fill = YELLOW
    ws["B5"].border = THIN
    ws["B5"].number_format = "0.000"
    ws["C5"] = "← seules entrées à changer au bureau"

    hdr = [
        "Section",
        "Station (m)",
        "S0 (-)",
        "y (m)",
        "WSE (m)",
        "Fond min (m)",
        "A (m²)",
        "P (m)",
        "R (m)",
        "V (m/s)",
        "Remarque",
    ]
    for j, h in enumerate(hdr, start=1):
        ws.cell(7, j, h)
    style_header(ws, 7, len(hdr))

    for i, (sid, snote, S0_real, computed, r) in enumerate(results, start=8):
        safe = sid.replace("+", "")
        sec_ref = f"Section_{safe}"
        if computed:
            remark = snote if r.ok else f"{snote} — {r.note}"
            vals = [
                sid,
                round(r.station_m, 3),
                round(S0_real, 5) if S0_real is not None else None,
                f"='{sec_ref}'!B10",
                f"='{sec_ref}'!B9",
                f"='{sec_ref}'!B9-'{sec_ref}'!B10",
                f"='{sec_ref}'!B11",
                f"='{sec_ref}'!B12",
                f"='{sec_ref}'!B13",
                f"='{sec_ref}'!B14",
                remark,
            ]
            row_fill = OK if r.ok else WARN
        else:
            vals = [
                sid,
                round(r.station_m, 3),
                round(S0_real, 5) if S0_real is not None else None,
                None,
                None,
                round(r.bed_elev, 3),
                None,
                None,
                None,
                None,
                snote,
            ]
            row_fill = SKIP

        for j, v in enumerate(vals, start=1):
            cell = ws.cell(i, j, v)
            cell.border = THIN
            if j in (2, 4, 5, 6, 7, 8, 9, 10):
                cell.number_format = "0.000"
            if j == 3 and isinstance(v, (int, float)):
                cell.number_format = "0.00000"
            if j == 3 and S0_real is not None and S0_real <= 0:
                cell.fill = SKIP
            elif j == 10 and computed:
                cell.fill = row_fill
            elif not computed and j in (4, 5, 7, 8, 9, 10):
                cell.fill = SKIP

    last = 7 + len(results)
    first_data = 8
    ws.cell(last + 2, 1, "V max")
    ws.cell(last + 2, 2, f"=MAX(J{first_data}:J{last})")
    ws.cell(last + 2, 2).fill = OK
    ws.cell(last + 2, 2).number_format = "0.000"
    ws.cell(last + 2, 3, "m/s")
    ws.cell(last + 3, 1, "V min")
    ws.cell(last + 3, 2, f"=MIN(J{first_data}:J{last})")
    ws.cell(last + 3, 2).number_format = "0.000"
    ws.cell(last + 3, 3, "m/s")
    ws.cell(last + 4, 1, "y max")
    ws.cell(last + 4, 2, f"=MAX(D{first_data}:D{last})")
    ws.cell(last + 4, 2).number_format = "0.000"
    ws.cell(last + 4, 3, "m")

    ws.cell(last + 6, 1, "Notes")
    ws.cell(last + 6, 1).font = Font(bold=True)
    notes = [
        "• Modifier Q (B4) ou n (B5) : le tableau se recalcule automatiquement (Excel).",
        "• Géométrie : coupes Distance–Élévation du levé.",
        "• S0 issu du profil longitudinal (valeur signée).",
        "• S0 ≤ 0 : profondeur normale non applicable (V non calculée).",
        "• S0 > 0 : Manning, V = Q/A à la profondeur normale.",
        "• Détail par coupe : onglets Section_* ; hypothèses : Notes_calcul.",
        "• Orange = pente adverse/nulle ; vert = V calculée.",
    ]
    for k, line in enumerate(notes):
        ws.cell(last + 7 + k, 1, line)
        ws.merge_cells(start_row=last + 7 + k, start_column=1, end_row=last + 7 + k, end_column=8)

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 14
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["K"].width = 42
    # Methode sheet
    wsM = wb.create_sheet("Methode")
    wsM["A1"] = "Méthode"
    wsM["A1"].font = Font(bold=True, size=13)
    for i, line in enumerate(
        [
            "",
            "Pour chaque section transversale :",
            "  1. S0 au profil longitudinal (station de la coupe).",
            "  2. Si S0 ≤ 0 : reporter S0 ; ne pas calculer yn ni V.",
            "  3. Si S0 > 0 : déterminer la cote d'eau telle que",
            "        Q = (1/n) · A · R^(2/3) · √S0",
            "     avec A et R calculés sur la polyligne de la coupe.",
            "  4. V = Q / A.",
            "",
            f"Q = {Q_DESIGN} m³/s ; n = {N_MANNING}",
            "Stations : 1+000 @ 0 m ; 2+000 @ 13,783 m ; 3+000 @ 23,663 m ; 4+000 @ 35,641 m ;",
            "           5+000 @ 45,838 m ; 6+000 @ 59,435 m ; 7+000 @ 71,973 m ; 8+000 @ 89,361 m.",
        ],
        start=2,
    ):
        wsM.cell(i, 1, line)
    wsM.column_dimensions["A"].width = 100

    # Longitudinal sheet
    wsL = wb.create_sheet("Longitudinal")
    wsL["A1"] = "Profil longitudinal"
    for j, h in enumerate(["Point", "Z fond (m)", "ΔL (m)", "S0 (-)", "Station (m)"], start=1):
        wsL.cell(3, j, h)
    style_header(wsL, 3, 5)
    for i, r in enumerate(long_rows, start=4):
        vals = [r["point"], r["z"], r["dL"], r["S0"], r["station"]]
        for j, v in enumerate(vals, start=1):
            cell = wsL.cell(i, j, v if v is not None else "—")
            cell.border = THIN
            if j in (2, 3, 5) and isinstance(v, (int, float)):
                cell.number_format = "0.000"
            if j == 4 and isinstance(v, (int, float)):
                cell.number_format = "0.00000"
                if v <= 0:
                    cell.fill = SKIP

    add_area_formula_sheets(wb, sections, results)

    wb.save(OUT_XLSX)
    print(f"Wrote {OUT_XLSX}")
    print(f"Clean CSV → {OUT_DIR}")


if __name__ == "__main__":
    main()
