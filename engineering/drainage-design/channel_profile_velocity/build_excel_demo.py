#!/usr/bin/env python3
"""Build Excel demo: inventé profile → velocity along channel for a chosen Q."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from profile_velocity import DEMO_DIR, run_profile, write_demo_csvs

OUT = Path(__file__).resolve().parent / "Profil_Vitesse_Demo_Invente.xlsx"

HEADER = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="0F5C5C")
YELLOW = PatternFill("solid", fgColor="FFF59D")
OK = PatternFill("solid", fgColor="D1FAE5")
WARN = PatternFill("solid", fgColor="FDE68A")
THIN = Border(
    left=Side(style="thin", color="CBD5E1"),
    right=Side(style="thin", color="CBD5E1"),
    top=Side(style="thin", color="CBD5E1"),
    bottom=Side(style="thin", color="CBD5E1"),
)


def style_header(ws, row: int, cols: int) -> None:
    for c in range(1, cols + 1):
        cell = ws.cell(row, c)
        cell.font = HEADER
        cell.fill = HEADER_FILL


def main() -> None:
    write_demo_csvs()
    Q = 2.0
    n = 0.035
    results = run_profile(Q, n)

    wb = Workbook()

    # --- Guide ---
    ws0 = wb.active
    ws0.title = "Methode"
    ws0["A1"] = "Démo inventée — vitesse le long d'un fossé (avant DXF AutoCAD)"
    ws0["A1"].font = Font(bold=True, size=14)
    lines = [
        "",
        "BUT: entrer un débit Q → obtenir profondeur y et vitesse V à chaque section.",
        "",
        "DONNÉES (inventées, dossier demo_data/):",
        "  • Longitudinal_demo.csv = station vs élévation du fond",
        "  • section_STA####.csv = coupe transversale (offset, elev) à cette station",
        "",
        "MÉTHODE (à chaque section):",
        "  1) Pente S0 depuis le profil longitudinal voisin",
        "  2) Essayer une cote d'eau WSE → calculer A et P sur la coupe",
        "  3) Q_manning = (1/n)·A·R^(2/3)·√S0   avec R=A/P",
        "  4) Ajuster WSE jusqu'à Q_manning ≈ Q  (profondeur normale)",
        "  5) V = Q / A",
        "",
        "FEUILLE Resultats: changez Q (jaune) puis relancez:  python build_excel_demo.py",
        "(Les résultats de section sont recalculés en Python — pas 100% formules Excel,",
        " car la résolution de y est itérative.)",
        "",
        "Quand vous aurez le DXF: remplacez les CSV dans demo_data/ ou incoming/",
        "par vos exports AutoCAD (mêmes colonnes) et relancez.",
    ]
    for i, line in enumerate(lines, start=2):
        ws0.cell(i, 1, line)
    ws0.column_dimensions["A"].width = 100

    # --- Inputs + results ---
    ws = wb.create_sheet("Resultats", 0)
    ws["A1"] = "Profil inventé — Manning le long du fossé"
    ws["A1"].font = Font(bold=True, size=13)
    ws["A2"] = "Éditez les jaunes, puis: python build_excel_demo.py"
    ws["A2"].font = Font(italic=True, color="64748B")

    ws["A4"] = "Q (m³/s)"
    ws["B4"] = Q
    ws["B4"].fill = YELLOW
    ws["C4"] = "débit imposé (steady)"
    ws["A5"] = "n Manning"
    ws["B5"] = n
    ws["B5"].fill = YELLOW
    ws["C5"] = "fossé herbeux / terre typ. 0.03–0.045"

    hdr = [
        "Station (m)",
        "Fond z_b (m)",
        "S0 (m/m)",
        "y profondeur (m)",
        "WSE (m)",
        "A (m²)",
        "P (m)",
        "R (m)",
        "V (m/s)",
        "Fr",
        "OK?",
        "Note",
    ]
    for j, h in enumerate(hdr, start=1):
        ws.cell(7, j, h)
    style_header(ws, 7, len(hdr))

    for i, r in enumerate(results, start=8):
        vals = [
            r.station_m,
            round(r.bed_elev, 3),
            round(r.S0, 5),
            round(r.y_m, 3),
            round(r.WSE, 3),
            round(r.A, 3),
            round(r.P, 3),
            round(r.R, 3),
            round(r.V, 3),
            round(r.Fr, 2),
            "Oui" if r.ok else "Non",
            r.note,
        ]
        for j, v in enumerate(vals, start=1):
            cell = ws.cell(i, j, v)
            cell.border = THIN
            if j == 9:  # V
                cell.fill = OK if r.ok else WARN

    last = 7 + len(results)
    ws.cell(last + 2, 1, "V max")
    ws.cell(last + 2, 2, round(max(r.V for r in results), 3))
    ws.cell(last + 2, 2).fill = OK
    ws.cell(last + 2, 3, "m/s")
    ws.cell(last + 3, 1, "y max")
    ws.cell(last + 3, 2, round(max(r.y_m for r in results), 3))
    ws.cell(last + 3, 3, "m")

    # --- Show demo longitudinal ---
    wsL = wb.create_sheet("Profil_longitudinal")
    wsL["A1"] = "Longitudinal_demo.csv (inventé)"
    wsL["A1"].font = Font(bold=True, size=12)
    wsL["A3"] = "station_m"
    wsL["B3"] = "elev_m (fond)"
    style_header(wsL, 3, 2)
    with (DEMO_DIR / "Longitudinal_demo.csv").open(encoding="utf-8") as f:
        next(f)
        for i, line in enumerate(f, start=4):
            s, z = line.strip().split(",")
            wsL.cell(i, 1, float(s)).border = THIN
            wsL.cell(i, 2, float(z)).border = THIN

    # --- One example section ---
    wsS = wb.create_sheet("Exemple_section")
    wsS["A1"] = "Exemple: section_STA0050.csv (station 50 m — point bas)"
    wsS["A1"].font = Font(bold=True, size=12)
    wsS["A3"] = "offset_m"
    wsS["B3"] = "elev_m"
    style_header(wsS, 3, 2)
    with (DEMO_DIR / "section_STA0050.csv").open(encoding="utf-8") as f:
        next(f)
        for i, line in enumerate(f, start=4):
            _sta, off, el = line.strip().split(",")
            wsS.cell(i, 1, float(off)).border = THIN
            wsS.cell(i, 2, float(el)).border = THIN
    wsS["A9"] = "Quand AutoCAD sera prêt: mêmes colonnes (station_m, offset_m, elev_m)."
    wsS["A9"].font = Font(italic=True, color="64748B")

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 14
    ws.column_dimensions["L"].width = 42

    wb.save(OUT)
    print(f"Wrote {OUT}")
    print(f"Q={Q} → V from {min(r.V for r in results):.2f} to {max(r.V for r in results):.2f} m/s")


if __name__ == "__main__":
    main()
